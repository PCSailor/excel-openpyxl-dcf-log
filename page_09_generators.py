#!/usr/bin/env python3
'''
'''
print('\nStart next file, \'page_09\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_09"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
columns = [(col, col+1) for col in range(2, 6, 2)]
center = Alignment(horizontal='center', vertical='center')
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# Local Variable
values = ['Fuel Storage Room 2', 'What fuel tank is selected', 'Leak detection panel', 'Room Temperature', 'Exhaust Fan', 'Total Gallons of Fuel', 'Generator Room', 'Air Compressor oil level', 'Generator 3', 'PRV 6 VFD is in Auto', 'FPCP Breaker is Closed', 'ATS-EB Breaker is Closed', 'LGS3-B03 ATS-MSB 3 Breaker is Closed', 'Generator 3 Coolant Temperature (˃100)', 'Day Tank Alarm_Gen3', 'Day Tank Alarm_Gen2', 'Battery Voltage', 'Battery Current', 'Oil Temperature (If  room temp., mark 105°F)', 'Generator 2', 'PRV 5 VFD is in Auto', 'LGS2-B03 ATS-MSB 2 Breaker is Closed', 'Generator 2 Coolant Temperature (˃100)', 'Battery Voltage', 'Battery Current', 'Oil Temperature (If room temp., mark 105°F)', 'Generator 1', 'PRV 4 VFD is in Auto', 'LGS1-B03 ATS-MSB 1 Breaker is Closed', 'Generator 1 Coolant temp. (˃100)', 'Day tank Alarms', 'Battery Voltage', 'Battery Current', 'Oil Temperature (If room temp., mark 105°F)', 'PRV 7 VFD is in Auto', 'Workshop', 'Fuel tank 1 Level', 'Fuel tank 2 Level', 'FCU- 9 VFD', 'E-Service', 'Fire pump breaker is Closed', 'Room Temperature', 'Notes']


def pg09_headers():
    # Print Options
    sheet.print_area = 'A1:I43' # Todo: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg09_merge():
    columns = [(col, col+1) for col in range(2, 10, 2)]
    # Merges 9 cells into 1 in 1 row
    for row in [1, 7, 9, 20, 27, 36, 40, 43]:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    # merge 2 cells into 1 in 1 row
    for row in [2, 3, 4, 5, 6, 11, 12, 13, 14, 15, 16, 17, 18, 19, 22, 23, 24, 25, 26, 29, 30, 31, 32, 33, 34, 37, 38, 39, 41, 42, 43]:
            for col1, col2 in columns:
                sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    # merge 8 cells in 1 column
    sheet.merge_cells(start_row=8, start_column=2, end_row=8, end_column=9)
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Dimensions_Columns & Rows
    sheet.column_dimensions['A'].width = 40.50
    for col in ['B', 'D', 'F', 'H']:
        sheet.column_dimensions[col].width = 6.75
    for col in ['C', 'E', 'G', 'I']:
        sheet.column_dimensions[col].width = 8.50
    rows = range(1, 46)
    for row in rows:
        sheet.row_dimensions[row].height = 15.00
    sheet.row_dimensions[43].height = 30.00
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Page Font
    rows = range(1, 55)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).font = Font(size = 10, i=False, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg09_namedstyle():
    # Styles
    sheet['A1'].style = 'rooms'
    sheet['A7'].style = 'rooms'
    sheet['A36'].style = 'rooms'
    sheet['A40'].style = 'rooms'
    sheet['A9'].style = 'subtitles'
    sheet['A20'].style = 'subtitles'
    sheet['A27'].style = 'subtitles'
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg09_cell_values():
    # Cell values
    def populate(sheet, col, row, values):
        row = int(row)
        for i, value in enumerate(values):
            sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list
    #
    sheet.cell(row=8, column=2).fill = PatternFill(fgColor='000000', fill_type = 'solid')
    sheet['A43'].alignment = leftTop
    sheet['A43'].font = Font(b=True)

def pg09_engineer_values():
    columns = range(2, 10, 1)
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    rowsCheck = [3, 5, 11, 12, 13, 15, 16, 22, 29, 31, 41]
    rowsT = [4, 14, 19, 23, 26, 30, 34, 42]
    rowsGals = [6, 37, 38]
    rowsHz = [10, 21, 28, 35, 39]
    rowsVdc = [17, 24, 32]
    rowsAmps = [18, 25, 33]
    rowsOnetwo = [2]
    rowsFault = []
    rowsBold = []
    # ✓ / X values
    for col in columns:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).value = '✓  /  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Gals
    for col in columns:
        for row in rowsGals:
            sheet.cell(row=row, column=col).value = 'Gals'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Temperature
    for col in columns:
        for row in rowsT:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '°F'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # ✓  /  X on Even
    for col in columnEven:
        for row in rowsHz:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = '✓ / X'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Hz on Odd
    for col in columnOdd:
        for row in rowsHz:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # vDC
    for col in columns:
        for row in rowsVdc:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'vDC'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Amps
    for col in columns:
        for row in rowsAmps:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'amps'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # 1 or 2
    for col in columns:
        for row in rowsOnetwo:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = '1 or 2'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    '''
    # Faults
    for col in columns:
        for row in rowsFault:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = 'Faults Yes / No'
            sheet.cell(row=row, column=col).font = Font(size=7, color='DCDCDC')
    # Bold text
    for col in columns:
        for row in rowsBold:
            sheet.cell(row=row, column=col).font = Font(bold=True, size=7)
    '''
    wb.save('Plymouth_Daily_Rounds.xlsx')


def pg09_colored_cells():
    columnsColor = range(1, 10, 1)
    rowsDkGrey = [1, 7, 36, 40]
    for col in columnsColor:
        for row in rowsDkGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    rowsLtGrey = [9, 20, 27]
    for col in columnsColor:
        for row in rowsLtGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Borders
    rows = range(1, 44)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')
