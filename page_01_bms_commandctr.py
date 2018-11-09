#! python3
'''
* Page 01 of Plymouth_Daily_Rounds.xlsx
* DCF Office & BMS Monitoring Computer
* Electrical Room Closets
* Command Center

* Bug: Row A1, columns 2, 3, 4, & 5 = double-border in red on top and bottom
* StretchGoal: increase Notes row height
* StretchGoal: border color sub-title rows in lighter grey
'''
print('\nStart next file, \'page_01\'')
'''
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
border = Border(top=double, left=thin, right=thin, bottom=double)
fill = PatternFill("solid", fgColor="DDDDDD")
font = Font(b=True, color="FF0000")

def style_range(sheet, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = sheet[cell_range.split(":")[0]]
    if alignment:
        sheet.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = sheet[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
'''

# imports
from openpyxl import load_workbook
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill, GradientFill
sheet = wb['Page_01']
print('Active sheet is ', sheet)
indexNumber = wb.worksheets.index(wb['Page_01'])
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
no_top_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'), 
                            top=Side(style='none'), 
                            bottom=Side(style='thin'))
no_bottom_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thick'))
# Local Variable
values = ['Mechanical / Chill Water Units Room'

]

def pg01_headers():
    # Local Variables
    # Print Options
    sheet.print_area = 'A1:E47' # TODO: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25 # Note: changed from .2 to .25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7 # Todo: Change is all sheets
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_merge():
    rows = range(1, 51)
    # 5 cells into 1 cell across 1 row
    for row in (1, 4, 5, 8, 20, 21, 28, 32, 33, 40, 44, 47, 48, 49, 59, 50):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    # Dimensions_Columns & Rows
    # Caution: No error is generated if the dimension value does not work
    sheet.column_dimensions['A'].width = 45.00
    for col in ['B', 'C', 'D', 'E']:
        sheet.column_dimensions[col].width = 14.00
    for row in rows:
        sheet.row_dimensions[row].height = 15.00
    sheet.row_dimensions[47].height = 20.00
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Page Font
    rows = range(1, 48)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).font = Font(size = 10, i=False, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')
    

def pg01_namedstyle():
    ''' NamedStyles set (mutable & used when need to apply formatting to different cells at once) '''
    # columnEven = [2, 4, 6, 8]
    # columnOdd = [3, 5, 7, 9]
    #
    headerrows = NamedStyle(name='headerrows')
    headerrows.font = Font(bold=True, sz=12)
    headerrows.alignment = center
    #
    rooms = NamedStyle(name='rooms')
    rooms.font = Font(b=True, sz=12)
    rooms.alignment = Alignment(horizontal='left', vertical='center')
    #
    subtitles = NamedStyle(name="subtitles")
    subtitles.font = Font(i=True, size=9)
    subtitles.alignment = left
    #
    rightAlign = NamedStyle(name='rightAlign')
    rightAlign.font = Font(b=True, i=True, sz=10)
    rightAlign.alignment = right
    wb.save('Plymouth_Daily_Rounds.xlsx')
    #
    # Header Rows
    sheet['A2'].style = rightAlign
    sheet['A3'].style = 'rightAlign'
    sheet['B3'].style = headerrows
    sheet['C3'].style = 'headerrows'
    sheet['D3'].style = 'headerrows'
    sheet['E3'].style = 'headerrows'
    # Room Divisions
    sheet['A4'].style = rooms
    sheet['A20'].style = 'rooms'
    sheet['A28'].style = 'rooms'
    sheet['A32'].style = 'rooms'
    # Subtitles
    sheet['A5'].style = subtitles
    sheet['A8'].style = 'subtitles'
    sheet['A21'].style = 'subtitles'
    sheet['A33'].style = 'subtitles'
    sheet['A40'].style = 'subtitles'
    sheet['A44'].style = 'subtitles'

    # Cell-specific adjustments 
    # A1
    a1 = sheet['A1']
    a1.font = Font(size=12, b=True, i=True, color='FF0000')
    a1.alignment = center
    a1.value = 'Note: When doing rounds be aware for unusual smells, sounds, sights, or anything not normal.'
    sheet['A47'].alignment = leftTop
    sheet['A47'].font = Font(b=True)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_cell_values():
    # Cell values
    ''' def populate(sheet, col, row, values):
            row = int(row)
            for i, value in enumerate(values):
                    sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list '''
    #
    sheet['A2'].value = 'Engineer Initials:     '
    sheet['A3'].value = 'Time of Round:     '
    sheet['B3'].value = '02:00'
    sheet['C3'].value = '08:00'
    sheet['D3'].value = '14:00'
    sheet['E3'].value = '20:00'
    # DCF Office
    sheet['A4'].value = 'DCF Office' # 
    sheet['A5'].value = 'Logs' # 
    sheet['A6'].value = 'Check Daily logs on Microsoft SharePoint' # Research: Add network location
    sheet['A7'].value = 'Check office white board for plant information'
    sheet['A8'].value = 'BMS'
    sheet['A9'].value = 'OAT (Over-the-Air Temperature)' # °F
    sheet['A10'].value = 'Wet Bulb' # °F
    sheet['A11'].value = 'Any BMS alarms? Note in comments below.'
    sheet['A12'].value = 'Chill Water Unit No. 1' # kW
    sheet['A13'].value = 'Chill Water Unit No. 2' # kW
    sheet['A14'].value = 'Chill Water Unit No. 3' # kW
    # Mechanical Screen
    sheet['A15'].value = 'Cooling Load_Mechanical Screen' # tR
    sheet['A16'].value = 'Tower Load_Mechanical Screen' # tR
    # Electrical Screen
    sheet['A17'].value = 'Total Power Usage_Electrical Screen' # kW
    sheet['A18'].value = 'DCF Power Usage_Electrical Screen' # kW
    sheet['A19'].value = 'PUE_Electrical Screen'
    sheet['A20'].value = 'Electrical Room_1st Floor' # 
    sheet['A21'].value = 'Fire Panel (Check for alarms on fire panel display)' # 
    sheet['A22'].value = 'Fire Alarm'
    sheet['A23'].value = 'Pre-alarm'
    sheet['A24'].value = 'Security'
    sheet['A25'].value = 'Supervisory'
    sheet['A26'].value = 'System Trouble'
    sheet['A27'].value = 'Other Event'
    sheet['A28'].value = 'Electrical Room_Lower Level' # 
    sheet['A29'].value = 'Room Temperature'
    sheet['A30'].value = 'EF 1' # Hz
    sheet['A31'].value = 'EF 2' # Hz
    sheet['A32'].value = 'Command Center'
    sheet['A33'].value = 'Fire Panel (Check for alarms on fire panel display)'
    sheet['A34'].value = 'Fire alarm'
    sheet['A35'].value = 'Pre-alarm'
    sheet['A36'].value = 'Security'
    sheet['A37'].value = 'Supervisory'
    sheet['A38'].value = 'System Trouble'
    sheet['A39'].value = 'Other Event'
    sheet['A40'].value = 'Vesda\'s' # 
    sheet['A41'].value = 'Server Room 1'
    sheet['A42'].value = 'Server Room 2'
    sheet['A43'].value = 'Server Room 3'
    sheet['A44'].value = 'Leak Detection' # 
    sheet['A45'].value = 'Server Rooms'
    sheet['A46'].value = 'PDU 13'
    sheet['A47'].value = 'Notes:' # StretchGoal: Increase row height, delete comment rows below
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_engineer_values():
    # Local Variables
    columns = [2, 3, 4, 5]
    rows = [11]
    rowsCheck = [6, 7, 22, 23, 24, 25, 26, 27, 29, 34, 35, 36, 37, 38, 39, 41, 42, 43, 45, 46]
    rowtR = [15, 16]
    rowT = [9, 10]
    rowkW = [12, 13, 14, 17, 18]
    rowsHZ = [30, 31]
    # Yes or No values
    for col in columns:
        for row in rows:
            sheet.cell(row=row, column=col).value = 'Yes   /   No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 9, color='696969')
    # ✓ X values
    for col in columns:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).value = '✓  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # tR
    for col in columns:
        for row in rowtR:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'tR'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Temperature
    for col in columns:
        for row in rowT:
            sheet.cell(row=row, column=col).value = '°F'
            sheet.cell(row=row, column=col).style = 'rightAlign'
            ''' sheet.cell(row=row, column=col).alignment = right '''
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Hz
    for col in columns:
        for row in rowsHZ:
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # kW
    rowsHZ = [18]
    for col in columns:
        for row in rowkW:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'kW'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_colored_cells():
    rowsLtGrey = [1, 5, 8, 21, 33, 40, 44]
    rowsDkGrey = [4, 20, 28, 32]
    columnsColor = [1, 2, 3, 4, 5]
    for col in columnsColor:
        for row in rowsDkGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    for col in columnsColor:
        for row in rowsLtGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    sheet.cell(row=2, column=1).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    sheet.cell(row=3, column=1).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    # sheet.cell(start_row=2, start_column=1, end_row=3, end_column=1).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Borders
    rows = range(1, 48)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')
