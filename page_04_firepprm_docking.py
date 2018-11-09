#! python3
'''
* East UPS Room
* Fire Pump Room
* Loading Dock Area
* Mechanical Room
'''
print('\nStart next file, \'page_04\'')
# imports
# from prettyprinter import pprint
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_04"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# Local Variable
values = ['** Ensure key is in locked position before touching STS display screen!! **', 'STS3A is on preferred source 1', 'STS3B is on preferred source 1', 'EF 4', 'EF 5', 'East Electrical Room Leak Detection', 'Tear off sticky mat for SR3 (East side)', 'Fire Pump/ Pre-Action Room', 'On 20:00 round only, \ncheck pre-action valves verifying open position, \ncheckmark zone if True):', '', '', '', '', '','Jockey pump controller in Auto', 'Fire pump controller in Auto', 'Fire pump is on Normal source power', 'System water pressure left side of fire pump controller (140 -150psi)', 'System Nitorgen PSI (inside the red cabinet) (+/- 25 psi)', 'Nitrogen Tank (Extra-Dry_2000PSI) (Replace at 200 psi)', 'Main building water meter (Total) readings (Top reading)', 'Is Building Main-Drain Water Pipe Leaking? (largest pipe at drain)', 'If drain pipe has water leaking, check the air-bleed-off-valve in the penthouse stairwell for leaks.', 'Loading Dock Area', 'Do we need to order salt? If yes, order & let Chief Engineer know.', 'Check brine level (should be at the indicating line).', 'HP LL- 5 Ok (Fan is ok, pipe sweating noticed, HP operation)', 'Mechanical / Chill Water Units Room', 'Cooling Twr. Supply  water meter reading.', 'Water Softener (Qty: 3)  Gallon Readings', '', '', 'Well meter reading', 'HP LL- 4 Ok (Fan is ok, pipe sweating noticed, HP operation)', 'CHWP #3' , 'CHWP #5', 'CHWP #2', 'CHWP #4', 'CHWP #1', 'CDW to CHW makeup' , 'CHW', 'CHW Filter PSI (23psi)', 'Bladder tank pressure (<30)' , 'CHW Lakos Bag filter', 'Notes:']
# pprint(values)

def pg04_headers():
    # Print Options
    sheet.print_area = 'A1:E45' # Todo: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_merge():
    # Merges 9 cells into 1 in 1 row
    for row in (1, 8, 23, 28, 45):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    # merge 3 cells in 1 row
    sheet.merge_cells(start_row=13, start_column=2, end_row=13, end_column=4)
    sheet.merge_cells(start_row=14, start_column=2, end_row=14, end_column=4)
    # merge 6 cells in 1 column
    sheet.merge_cells(start_row=9, start_column=1, end_row=14, end_column=1)
    sheet.merge_cells(start_row=30, start_column=1, end_row=32, end_column=1) # DELETE??
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Dimensions_Columns & Rows
    sheet.column_dimensions['A'].width = 60.00
    for col in ['B', 'C', 'D', 'E']:
        sheet.column_dimensions[col].width = 10.00
    rows = range(1, 50)
    for row in rows:
        sheet.row_dimensions[row].height = 15.00
    sheet.row_dimensions[45].height = 40.00 # Notes
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Page Font
    rows = range(1, 45)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).font = Font(size = 10, i=False, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_namedstyle():
    # Styles
    sheet['A8'].style = 'rooms'
    sheet['A24'].style = 'rooms'
    sheet['A28'].style = 'rooms'
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_cell_values():
    # Cell values
    def populate(sheet, col, row, values):
        row = int(row)
        for i, value in enumerate(values):
            sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list
    #
    sheet['B9'].value = 'Zone 1'
    sheet['B10'].value = 'Zone 2'
    sheet['B11'].value = 'Zone 3'
    sheet['B12'].value = 'Zone 4'
    sheet['B13'].value = 'Wet system level 1-4 '
    sheet['B14'].value = 'Wet system level 0 (Corridors)'
    sheet['D9'].value = 'Zone 5'
    sheet['D10'].value = 'Zone 6'
    sheet['D11'].value = 'Zone 7'
    #
    sheet['B30'].value = 'Softener#1'
    sheet['C30'].value = 'Softener#1'
    sheet['D30'].value = 'Softener#1'
    sheet['E30'].value = 'Softener#1'
    sheet['B31'].value = 'Softener#2' 
    sheet['C31'].value = 'Softener#2' 
    sheet['D31'].value = 'Softener#2' 
    sheet['E31'].value = 'Softener#2' 
    sheet['B32'].value = 'Softener#3'
    sheet['C32'].value = 'Softener#3'
    sheet['D32'].value = 'Softener#3'
    sheet['E32'].value = 'Softener#3'
    #
    sheet['A1'].font = Font(size=12, b=True, i=True, color='FF0000')
    sheet['A1'].alignment = center
    sheet['A9'].alignment = center
    sheet['A23'].font = Font(size=10, b=True, i=True, color='FF0000')
    sheet['A27'].alignment = left
    sheet['A30'].alignment = left
    sheet['A45'].alignment = leftTop
    sheet['A45'].font = Font(b=True)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_engineer_values():
    columns = range(2, 6, 1)
    # columnOdd = [3, 5] # 9, 10, 11, 12, 13, 14
    rowYes = [7, 15, 16, 17, 22, 25]
    rowsCheck = [2, 3, 6, 26, 27, 34, 38]
    rowsHZ = [4, 5]
    rowsDp = [35, 36, 37, 38, 39, 40, 41, 44]
    rowPsi = [42, 43]
    rowSoft = [30, 31, 32]
    # Yes or No values
    for col in columns:
        for row in rowYes:
            sheet.cell(row=row, column=col).value = 'Yes  /  No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 9, color='696969')
    # ✓ / X values
    for col in columns:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).value = '✓  /  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
        # Hz
    for col in columns:
        for row in rowsHZ:
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # DP
    for col in columns:
        for row in rowsDp:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'DP'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')    
    # PSI
    for col in columns:
        for row in rowPsi:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'psi'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Softener
    for col in columns:
        for row in rowSoft:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_colored_cells():
    rowsDkGrey = [8, 24, 28]
    columnsColor = range(1, 6, 1)
    for col in columnsColor:
        for row in rowsDkGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    ''' rowsLtGrey = []
    for col in columnsColor:
        for row in rowsLtGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid') '''
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Borders
    rows = range(1, 46)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')

