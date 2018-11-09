#!/usr/bin/env python3
'''
* 
wb.save('Plymouth_Daily_Rounds.xlsx')
'''
print('\nStart next file, \'page_06\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_06"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
columns = [(col, col+1) for col in range(2, 6, 2)]
center = Alignment(horizontal='center', vertical='center')
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# Local Variable
values = ['Mechanical / Chill Water Units Room continued'

]

def pg06_headers():
    # Print Options
    sheet.print_area = 'A1:E52' # Todo: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg06_merge():
    columns = [(col, col+1) for col in range(2, 6, 2)]
    # Merges 5 cells into 1 in 1 row
    for row in (6, 22, 37, 52):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    # merge 2 cells into 1 in 1 row
    for row in []:
            for col1, col2 in columns:
                sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Dimensions_Columns & Rows
    sheet.column_dimensions['A'].width = 60.00
    for col in ['B', 'C', 'D', 'E']:
        sheet.column_dimensions[col].width = 10
    rows = range(1, 55)
    for row in rows:
        sheet.row_dimensions[row].height = 13.75
    sheet.row_dimensions[52].height = 24.00
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Page Font
    rows = range(1, 55)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).font = Font(size = 10, i=False, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg06_namedstyle():
    # Styles
    sheet['A6'].style = 'subtitles' # question: local vs global vars?
    sheet['A22'].style = 'subtitles'
    sheet['A37'].style = 'subtitles'
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg06_cell_values():
    # Cell values
    ''' def populate(sheet, col, row, values):
            row = int(row)
            for i, value in enumerate(values):
                    sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list '''
    #
    sheet['A1'].value = 'TWP 1 DP'
    sheet['A2'].value = 'TWP 6 DP'
    sheet['A3'].value = 'East Sump Level'
    sheet['A4'].value = 'Lakos cond. Water(7 PSID)'
    sheet['A5'].value = 'ATS-HPB-H Load on Normal'
    sheet['A6'].value = 'Chiller #1' # Room - Chiller #1
    sheet['A7'].value = 'In Recycle mode?? If so, fill out bold items'
    sheet['A8'].value = 'Is online?? If not, fill out only the last 3 bold items'
    sheet['A9'].value = 'Evap PSID 40” – 70”'
    sheet['A10'].value = 'Cond PSID 40” – 70”'
    sheet['A11'].value = 'CHW In temp. ≤ 65 deg.'
    sheet['A12'].value = 'CHW Out temp. 44 – 57deg.'
    sheet['A13'].value = 'Evap Ref.  43 – 56 deg.'
    sheet['A14'].value = 'CDW In Temp. 60- 90 deg.'
    sheet['A15'].value = 'CDW Out Temp. <CWP + 10 deg.'
    sheet['A16'].value = 'Cond. Ref. (65 – 115deg.)'
    sheet['A17'].value = 'Oil Pressure (25- 35 Psi)'
    sheet['A18'].value = 'Oil Temp. (70- 130 deg.)'
    sheet['A19'].value = 'Running in CCN mode'
    sheet['A20'].value = 'Manual/Auto switch in Auto'
    sheet['A21'].value = 'HP LL- 2  Ok (Fan is ok, pipe sweating noticed, HP operation)'
    sheet['A22'].value = 'Chiller #2' # Room - Chiller #2
    sheet['A23'].value = 'In Recycle mode?? If so, fill out bold items'
    sheet['A24'].value = 'Is online?? If not, fill out only the last 3 bold items'
    sheet['A25'].value = 'Evap PSID 40” – 70”'
    sheet['A26'].value = 'Cond PSID 40” – 70”'
    sheet['A27'].value = 'CHW In temp. ≤ 65 deg.'
    sheet['A28'].value = 'CHW Out temp. 44 – 57deg.'
    sheet['A29'].value = 'Evap Ref.  43 – 56 deg.'
    sheet['A30'].value = 'CDW In Temp. 60- 90 deg.'
    sheet['A31'].value = 'CDW Out Temp. <CWP + 10 deg.'
    sheet['A32'].value = 'Cond. Ref. (65 – 115deg.)'
    sheet['A33'].value = 'Oil Pressure (25- 35 Psi)'
    sheet['A34'].value = 'Oil Temp. (70- 130 deg.)'
    sheet['A35'].value = 'Running in CCN mode'
    sheet['A36'].value = 'Manual/Auto switch in Auto'
    sheet['A37'].value = 'Chiller #3' # Room - Chiller #3'
    sheet['A38'].value = 'In Recycle mode?? If so, fill out bold items'
    sheet['A39'].value = 'Is online?? If not, fill out only the last 3 bold items'
    sheet['A40'].value = 'Evap PSID 40” – 70”'
    sheet['A41'].value = 'Cond PSID 40” – 70”'
    sheet['A42'].value = 'CHW In temp. ≤ 65 deg.'
    sheet['A43'].value = 'CHW Out temp. 44 – 57deg.'
    sheet['A44'].value = 'Evap Ref.  43 – 56 deg.'
    sheet['A45'].value = 'CDW In Temp. 60- 90 deg.'
    sheet['A46'].value = 'CDW Out Temp. <CWP + 10 deg.'
    sheet['A47'].value = 'Cond. Ref. (65 – 115deg.)'
    sheet['A48'].value = 'Oil Pressure (25- 35 Psi)'
    sheet['A49'].value = 'Oil Temp. (70- 130 deg.)'
    sheet['A50'].value = 'Running in CCN mode'
    sheet['A51'].value = 'Manual/Auto switch in Auto'
    sheet['A52'].value = 'Notes'
    sheet['A52'].alignment = leftTop
    sheet['A52'].font = Font(b=True)

def pg06_engineer_values():
    columns = range(2, 6, 1)
    rowsYes = [7, 8, 19, 20, 23, 24, 35, 36, 38, 39, 50, 51]
    rowsCheck = [5, 21]
    rowsDp = [1, 2, 4, 9, 10, 25, 26, 40, 41]
    rowsT = [11, 12, 13, 14, 15, 16, 18, 27, 28, 29, 30, 31, 32, 34, 42, 43, 44, 45, 46, 47, 49]
    rowsPsi = [17, 33, 48]
    rowsInch = [3]
    rowsBold = [7, 9, 18, 19, 20, 23, 25, 34, 35, 36, 38, 40, 49, 50, 51]
    # Yes or No values
    for col in columns:
        for row in rowsYes:
            sheet.cell(row=row, column=col).value = 'Yes  /  No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 9, color='696969')
    # ✓ / X values
    for col in columns:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).value = '✓  /  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Temperature
    for col in columns:
        for row in rowsT:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '°F'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # DP
    for col in columns:
        for row in rowsDp:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'DP'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # PSI
    for col in columns:
        for row in rowsPsi:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'PSI'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Inches of Water
    for col in columns:
        for row in rowsInch:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'inH2O'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Bold text
    for col in columns:
        for row in rowsBold:
            sheet.cell(row=row, column=col).font = Font(bold=True, size=7)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg06_colored_cells():
    columnsColor = range(1, 6, 1)
    rowsLtGrey = [6, 22, 37]
    for col in columnsColor:
        for row in rowsLtGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Borders
    rows = range(1, 53)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')