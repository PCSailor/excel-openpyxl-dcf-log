'''
* Page 01 of Plymouth_Daily_Rounds.xlsx
* DCF Office & BMS Monitoring Computer
* Electrical Room Closets
* Command Center
'''
print('\nStart next file, \'page_01_bms_commandctr.py\'')
# imports
# import openpyxl
from openpyxl import load_workbook
# from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill, NamedStyle, Color, colors
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill# , GradientFill
sheet = wb['Page_01']
print('Worksheet list:', wb.sheetnames) # 'Plymouth_Daily_Rounds', 'Page_11'
indexNumber = wb.worksheets.index(wb['Page_01'])
print('\'Page_01\' index number = ', indexNumber) # = 0
indexNumber = wb.worksheets.index(wb['Page_11'])
print('\'Page_11\' index number = ', indexNumber) # 1
print('Active sheet is ', sheet, '\n')
wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_start():
    # sheet['A1'].value = 'Note: When doing rounds be aware for unusual smells, sounds, sights, or anything not normal.' 
    print('startup complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_headers():
    # Local Variables
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='bottom')
    # Print Options
    sheet.print_area = 'A1:E47' # TODO: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000" # 
    print('headers complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_merge():
    # Local Variables
    rows = range(1, 51)
    # 5 cells into 1 cell across 1 row
    for row in (1, 4, 5, 8, 20, 21, 28, 32, 33, 40, 44, 47, 48, 49, 59, 50):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    #
    # Column and Row Dimensions
    # Caution: No error is generated if the dimension value does not work
    sheet.column_dimensions['A'].width = 45.00
    for col in ['B', 'C', 'D', 'E']:
        sheet.column_dimensions[col].width = 14.00
    for row in rows:
        sheet.row_dimensions[row].height = 15.00
    sheet.row_dimensions[47].height = 21.00
    print('merge complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_namedstyle():
    ''' NamedStyles set (mutable & used when need to apply formatting to different cells at once) '''
    print('pg01_namedstyle is run')
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='bottom')
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), 
                        right=Side(style='thick'), 
                        top=Side(style='thick'), 
                        bottom=Side(style='thick'))
    no_top_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'), 
                            top=Side(style='none'), 
                            bottom=Side(style='thin'))
    no_bottom_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thick'))
    no_border = Border(left=Side(style='none'),
                            right=Side(style='none'), 
                            top=Side(style='none'), 
                            bottom=Side(style='none'))
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    #
    headerrows = NamedStyle(name='headerrows')
    headerrows.font = Font(bold=True, underline='none', sz=12)
    headerrows.alignment = center
    #
    rooms = NamedStyle(name='rooms')
    rooms.font = Font(b=True, sz=12)
    rooms.alignment = Alignment(horizontal='left', vertical='center')
    #
    subtitles = NamedStyle(name="subtitles")
    subtitles.font = Font(i=True, size=9)
    #
    rightAlign = NamedStyle(name='rightAlign')
    rightAlign.font = Font(b=True, i=True, sz=10)
    rightAlign.alignment = right
    #
    print('namedstyle complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')
    #
    # A1
    a1 = sheet['A1']
    a1.style = rooms
    a1.font = Font(size=12, b=False, i=True, color='FF0000')
    a1.alignment = center
    a1.value = 'Note: When doing rounds be aware for unusual smells, sounds, sights, or anything not normal.'
    #
    # Header Rows
    sheet['A2'].style = rightAlign
    sheet['A3'].style = 'rightAlign'
    sheet['B3'].style = headerrows
    sheet['C3'].style = 'headerrows'
    sheet['D3'].style = 'headerrows'
    sheet['E3'].style = 'headerrows'
    # Room Divisions
    sheet['A4'].style = rooms
    sheet['A20'].style = 'rooms'
    sheet['A28'].style = 'rooms'
    sheet['A32'].style = 'rooms'
    sheet['A47'].style = 'rooms'
    # Subtitles
    sheet['A5'].style = subtitles
    sheet['A8'].style = 'subtitles'
    sheet['A21'].style = 'subtitles'
    sheet['A33'].style = 'subtitles'
    sheet['A40'].style = 'subtitles'
    sheet['A44'].style = 'subtitles'
    #
    # Set Borders
    ''' The merged cell behaves similar to other cell ojects. 
    Its value and format is defined in its top-left cell. 
    In order to change the border of the whole merged cell, change the border of its top-left cell. '''
    rows = range(1, 49)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = no_border
    # StretchGoal: Need Thick border around page doc
    print('styles complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_cell_values():
    # Local Variables
    # Cell values
    sheet['A2'].value = 'Engineer Initials:     '
    sheet['A3'].value = 'Time of Round:     '
    sheet['B3'].value = '02:00'
    sheet['C3'].value = '08:00'
    sheet['D3'].value = '14:00'
    sheet['E3'].value = '20:00'
    # DCF Office
    sheet['A4'].value = 'DCF Office' # 
    sheet['A5'].value = 'Logs' # 
    sheet['A6'].value = 'Check Daily logs on Microsoft SharePoint' # Research: Add network location
    sheet['A7'].value = 'Check office white board for plant information'
    sheet['A8'].value = 'BMS' # 
    sheet['A9'].value = 'OAT (Over-the-Air Temperature)'
    sheet['A10'].value = 'Wet Bulb'
    sheet['A11'].value = 'Any BMS alarms? Note in comments below.'
    sheet['A12'].value = 'Chill Water Unit No. 1'
    sheet['A13'].value = 'Chill Water Unit No. 2'
    sheet['A14'].value = 'Chill Water Unit No. 3'
    # sheet['A14'].value = 'Mechanical Screen' # 
    sheet['A15'].value = 'Cooling Load_Mechanical Screen'
    sheet['A16'].value = 'Tower Load_Mechanical Screen'
    # sheet['A17'].value = 'Electrical Load_Electrical Screen'
    sheet['A17'].value = 'Total Power Usage_Electrical Screen'
    sheet['A18'].value = 'DCF Power Usage_Electrical Screen'
    sheet['A19'].value = 'PUE_Electrical Screen'
    sheet['A20'].value = 'Electrical Room_1st Floor' # 
    sheet['A21'].value = 'Fire Panel (Check for alarms on fire panel display)' # 
    sheet['A22'].value = 'Fire Alarm'
    sheet['A23'].value = 'Pre-alarm'
    sheet['A24'].value = 'Security'
    sheet['A25'].value = 'Supervisory'
    sheet['A26'].value = 'System Trouble'
    sheet['A27'].value = 'Other Event'
    sheet['A28'].value = 'Electrical Room_Lower Level' # 
    sheet['A29'].value = 'Room Temperature'
    sheet['A30'].value = 'EF 1'
    sheet['A31'].value = 'EF 2'
    sheet['A32'].value = 'Command Center' # 
    sheet['A33'].value = 'Fire Panel (Check for alarms on fire panel display)' # 
    # sheet['A35'].value = ':' # 
    sheet['A34'].value = 'Fire alarm'
    sheet['A35'].value = 'Pre-alarm'
    sheet['A36'].value = 'Security'
    sheet['A37'].value = 'Supervisory'
    sheet['A38'].value = 'System Trouble'
    sheet['A39'].value = 'Other Event'
    sheet['A40'].value = 'Vesda\'s' # 
    sheet['A41'].value = 'Server Room 1'
    sheet['A42'].value = 'Server Room 2'
    sheet['A43'].value = 'Server Room 3'
    sheet['A44'].value = 'Leak Detection' # 
    sheet['A45'].value = 'Server Rooms'
    sheet['A46'].value = 'PDU 13'
    sheet['A47'].value = 'Notes:' # StretchGoal: Increase row height, delete comment rows below
    print('cell_values complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_engineer_values():
    # Local Variables
    columns = [2, 3, 4, 5]
    rows = [11]
    rowsCheck = [6, 7, 22, 23, 24, 25, 26, 27, 29, 34, 35, 36, 37, 38, 39, 41, 42, 43, 45, 46]
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='bottom')
    # Yes or No values
    for col in columns:
        for row in rows:
            sheet.cell(row=row, column=col).value = 'Yes   /   No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 9, i=True, color='000000')
    # ✓ X values
    for col in columns:
        for row in rowsCheck:
            # print(col, row)
            sheet.cell(row=row, column=col).value = '✓  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=9, color='DCDCDC')
    print('engineer_values complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg01_colored_cells():
    # rowsColor = [1, 2, 3, 4, 24, 41]
    rowsColor = [4, 20, 28, 32]
    columnsColor = [1, 2, 3, 4, 5]
    for col in columnsColor:
        for row in rowsColor:
            # print(col, row)
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    print('colored_cells complete on ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')