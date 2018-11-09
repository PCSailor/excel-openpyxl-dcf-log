#! python3
'''
* Mechanical Room
'''
print('\nStart next file, \'page_05\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_05"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
columns = [(col, col+1) for col in range(2, 9, 2)]
center = Alignment(horizontal='center', vertical='center')
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# Local Variable
values = ['Mechanical / Chill Water Units Room continued', 'Condenser Supply_East (Normal = 68 to 85)', 'CWP-6 VFD', 'CWP-1 VFD', 'CWP-4 VFD', 'CWP-3 VFD', 'CDWF  VFD', 'CWP-2 VFD', 'CWP-5 VFD', 'TWR Fan- 6 VFD', 'TWR Fan- 5 VFD', 'CHWR Header Temp East', 'CHWR Temp (Bypass) East', 'Lakos Separator (6psi)', 'CHWS Temp East', 'CHWP #3 VFD', 'Well VFD', 'CHWP #2 VFD', 'CHWP #4 VFD', 'CHWP #1 VFD', 'CHWP #5 VFD', 'EF #6 VFD', 'Core Pump #1 VFD', 'Core Pump #2 VFD', 'HP LL- 3 Ok (Fan is ok, pipe sweating noticed, HP operation)', 'Core Pump #2 (15 - 20 PSID)', 'Core Pump #1 (15 - 20 PSID)', 'Condenser Supply_West (Normal = 68 to 85)', 'Chemical tanks level (above the order lines)', 'Nalco controller', 'Coupon Rack flow is between 4 – 6 GPM (Clean Strainer)', 'Tower #4 VFD', 'Tower #3 VFD', 'Tower #2 VFD', 'Tower #1 VFD', 'Core Filter PSI (38 to 45 psi)', 'HEX Core Water Temp In', 'HEX Core Water Temp Out', 'HEX Cond Water Temp In', 'HEX Cond Water Temp Out', 'HEX Cond DP (10psi min.)', 'HEX Core DP (150” PSID)', 'West sump level ', 'TWP 5 DP', 'TWP 2 DP', 'TWP 3 DP', 'TWP 4 DP', 'Notes:']


def pg05_headers():
    # Print Options
    sheet.print_area = 'A1:I48' # Todo: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg05_merge():
        columns = [(col, col+1) for col in range(2, 9, 2)]
        # Merges 9 cells into 1 in 1 row
        for row in (1, 48):
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        # merge 2 cells into 1 in 1 row
        for row in [1, 11, 12, 13, 14, 24, 25, 26, 27, 28, 29, 30, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47]:
                for col1, col2 in columns:
                        sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
        wb.save('Plymouth_Daily_Rounds.xlsx')

        # Dimensions_Columns & Rows
        rows = range(1, 46)
        sheet.column_dimensions['A'].width = 45.00
        for col in ['B', 'D', 'F', 'H']:
                sheet.column_dimensions[col].width = 5.00
        for col in ['C', 'E', 'G', 'I']:
                sheet.column_dimensions[col].width = 9.50
        for row in rows:
                sheet.row_dimensions[row].height = 14.75
        sheet.row_dimensions[48].height = 20.00
        wb.save('Plymouth_Daily_Rounds.xlsx')

        # Page Font
        rows = range(1, 50)
        columns = range(1, 10)
        for row in rows:
                for col in columns:
                        sheet.cell(row, col).font = Font(size = 10, i=False, color='000000')
        wb.save('Plymouth_Daily_Rounds.xlsx')

def pg05_namedstyle():
        # Styles
        sheet['A1'].style = 'rooms'
        wb.save('Plymouth_Daily_Rounds.xlsx')

def pg05_cell_values():
        # Cell values
        def populate(sheet, col, row, values):
                row = int(row)
                for i, value in enumerate(values):
                        sheet['{}{}'.format(col, row+i)].value = value
        populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list
        
        sheet['A48'].alignment = leftTop
        sheet['A48'].font = Font(b=True)
        wb.save('Plymouth_Daily_Rounds.xlsx')

def pg05_engineer_values():
    # Local Variables
    columns = range(2, 9, 1)
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    rowsCheck = [2, 3, 4, 5, 6, 7, 8, 9, 10, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 28, 29, 31, 32, 33, 34]
    rowsHz = [2, 3, 4, 5, 6, 7, 8, 9, 10, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 31, 32, 33, 34]
    rowsDp = [13, 25, 26, 43, 44, 45, 46]
    rowsT = [1, 11, 12, 14, 27, 36, 37, 38, 39]
    rowsPsi = [35, 40, 41]
    rowsGpm = [30]
    rowsinH2O = [42]
    # ✓ / X values
    for col in columnEven:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).value = '✓  /  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
        # Hz
    for col in columnOdd:
        for row in rowsHz:
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Temperature
    for col in columnEven:
        for row in rowsT:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '°F'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')  
    # DP
    for col in columnEven:
        for row in rowsDp:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'DP'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')          
    # PSI
    for col in columnEven:
        for row in rowsPsi:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'PSI'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # GPM
    for col in columnEven:
        for row in rowsGpm:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'GPM'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Inches Water
    for col in columnEven:
        for row in rowsinH2O:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'inHg'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg05_colored_cells():
        rowsDkGrey = [1]
        columnsColor = range(1, 6, 1)
        for col in columnsColor:
                for row in rowsDkGrey:
                        sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
        wb.save('Plymouth_Daily_Rounds.xlsx')

         # Borders
        rows = range(1, 49)
        columns = range(1, 10)
        for row in rows:
                for col in columns:
                        sheet.cell(row, col).border = thin_border
        wb.save('Plymouth_Daily_Rounds.xlsx')