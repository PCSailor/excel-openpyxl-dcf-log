#!/usr/bin/env python3
'''
* 
wb.save('Plymouth_Daily_Rounds.xlsx')
'''
print('\nStart next file, \'page_08\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_08"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
columns = [(col, col+1) for col in range(2, 6, 2)]
rows = range(1, 47)
center = Alignment(horizontal='center', vertical='center')
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# Local Variable
''' values = ['Mechanical / Chill Water Units Room continued', 'DP 3 Breakers', 'DP3-B08 SPD All 3 Green lights (Protected)'
    sheet['A4'].value = 'DP3-B12 CHILLER 3 Breaker is Closed'
    sheet['A5'].value = 'DP3-B13 ATS-HPC-C Breaker is Closed'
    sheet['A6'].value = 'DP3-B14 ATS-HPC-H Breaker is Closed'
    sheet['A7'].value = 'DP3-B15 MUA 4 Breaker is Closed'
    sheet['A8'].value = 'DP3-B17 ATS-HPD-H Breaker is Closed'
    sheet['A9'].value = 'DP3-B18 ATS-HPC1-C Breaker is Closed'
    sheet['A10'].value = 'DP3-B19 ATS-HPB1-C Breaker is Closed'
    sheet['A11'].value = 'DP3-B20 Spare Breaker is Closed'
    # sheet['A12'].value = 'MAU 3 VFD' # Subtitles
    sheet['A12'].value = 'MAU 3 VFD_Alarms'
    # sheet['A14'].value = 'Remote Radiators' # Subtitles
    sheet['A13'].value = 'Remote Radiator RR1 VFD in Auto'
    sheet['A14'].value = 'Remote Radiator RR2 VFD in Auto'
    sheet['A15'].value = 'Remote Radiator RR3 VFD in Auto'
    # sheet['A18'].value = 'ATS\'s' # Subtitles
    sheet['A16'].value = 'ATS-HPD-H Load on Normal'
    sheet['A17'].value = 'ATS-HPA1-C Load on Normal'
    sheet['A18'].value = 'ATS-HPB1-C Load on Normal'
    sheet['A19'].value = 'ATS-HPC1-C Load on Normal'
    sheet['A20'].value = 'MSB 2' # Subtitles
    sheet['A21'].value = 'MSB2-B12 CI2 Breaker is Closed'
    sheet['A22'].value = 'MSB2-B13 DP2 Breaker is Closed'
    sheet['A23'].value = 'MSB2-B14 DP1 Temp Feed Breaker is racked out'
    sheet['A24'].value = 'Eaton Xpert meter Events light off'
    sheet['A25'].value = 'MSB2-B01 Main Breaker is Closed'
    sheet['A26'].value = 'MSB2-B08 SPD All 3 Green lights (Protected)'
    sheet['A27'].value = 'RPTCS (S1 lights are on)'
    sheet['A28'].value = 'Mode Switch is in the `Closed Transition position'
    sheet['A29'].value = 'Tower Yard' # Rooms
    sheet['A30'].value = 'Check the Towers for any Leaks and noises (Bearings, Belts ect.)'
    sheet['A30'].font = Font(size=7, b=False, i=True, color='FF0000')
    sheet['A30'].alignment = center
    # sheet['A30'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet['A31'].value = 'Tower 1'
    sheet['A32'].value = 'Tower 2'
    sheet['A33'].value = 'Tower 3'
    sheet['A34'].value = 'Tower 4'
    sheet['A35'].value = 'Tower 5'
    sheet['A36'].value = 'Tower 6'
    sheet['A37'].value = 'MUA 3 Status and Coils are clean and OK'
    sheet['A38'].value = 'Fuel Storage Room 1' # Rooms
    sheet['A39'].value = 'Total Gallons of Fuel'
    sheet['A40'].value = 'Leak detection panel'
    sheet['A41'].value = 'Room Temperature'
    sheet['A42'].value = 'EF'
    sheet['A43'].value = '1000 Gallon Day Tank Level'
    sheet['A44'].value = 'Alarms'
    sheet['A45'].value = 'Notes'

]
'''

def pg08_headers():
    # Print Options
    sheet.print_area = 'A1:I45' # Todo: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg08_merge():
    # Merges 9 cells into 1 in 1 row
    for row in (1, 2, 20, 29, 38, 45):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    columns = [(col, col+1) for col in range(2, 9, 2)]
    # merge 2 cells into 1 in 1 row
    for row in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 27, 28, 37, 39, 40, 41, 42, 43, 44]:
            for col1, col2 in columns:
                sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    # merge 8 cells into 1 in 1 row
    for row in [23]:
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Dimensions_Columns & Rows
    sheet.column_dimensions['A'].width = 49.75
    for col in ['B', 'D', 'F', 'H']:
        sheet.column_dimensions[col].width = 6.00
    for col in ['C', 'E', 'G', 'I']:
        sheet.column_dimensions[col].width = 6.00
    rows = range(1, 50)
    for row in rows:
        sheet.row_dimensions[row].height = 15.00
    sheet.row_dimensions[45].height = 30.00
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Page Font
    rows = range(1, 50)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).font = Font(size = 10, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Rows 30-36, column A centering
    rows = (30, 31, 32, 33, 34, 35, 36)
    for row in rows:
        sheet.cell(row=row, column=1).alignment = center

def pg08_namedstyle():
    # Styles
    sheet['A1'].style = 'rooms'
    sheet['A29'].style = 'rooms'
    sheet['A38'].style = 'rooms'
    sheet['A2'].style = 'subtitles'
    sheet['A20'].style = 'subtitles'
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg08_cell_values():
    # Cell values
    ''' def populate(sheet, col, row, values):
            row = int(row)
            for i, value in enumerate(values):
                    sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list '''
    #
    sheet['A1'].value = 'Mechanical / Chill Water Units Room continued'
    sheet['A1'].font = Font(size=8, b=True, i=True)
    sheet['A2'].value = 'DP 3 Breakers' # Subtitles
    sheet['A3'].value = 'DP3-B08 SPD All 3 Green lights (Protected)'
    sheet['A4'].value = 'DP3-B12 CHILLER 3 Breaker is Closed'
    sheet['A5'].value = 'DP3-B13 ATS-HPC-C Breaker is Closed'
    sheet['A6'].value = 'DP3-B14 ATS-HPC-H Breaker is Closed'
    sheet['A7'].value = 'DP3-B15 MUA 4 Breaker is Closed'
    sheet['A8'].value = 'DP3-B17 ATS-HPD-H Breaker is Closed'
    sheet['A9'].value = 'DP3-B18 ATS-HPC1-C Breaker is Closed'
    sheet['A10'].value = 'DP3-B19 ATS-HPB1-C Breaker is Closed'
    sheet['A11'].value = 'DP3-B20 Spare Breaker is Closed'
    # sheet['A12'].value = 'MAU 3 VFD' # Subtitles
    sheet['A12'].value = 'MAU 3 VFD_Alarms'
    # sheet['A14'].value = 'Remote Radiators' # Subtitles
    sheet['A13'].value = 'Remote Radiator RR1 VFD in Auto'
    sheet['A14'].value = 'Remote Radiator RR2 VFD in Auto'
    sheet['A15'].value = 'Remote Radiator RR3 VFD in Auto'
    # sheet['A18'].value = 'ATS\'s' # Subtitles
    sheet['A16'].value = 'ATS-HPD-H Load on Normal'
    sheet['A17'].value = 'ATS-HPA1-C Load on Normal'
    sheet['A18'].value = 'ATS-HPB1-C Load on Normal'
    sheet['A19'].value = 'ATS-HPC1-C Load on Normal'
    sheet['A20'].value = 'MSB 2' # Subtitles
    sheet['A21'].value = 'MSB2-B12 CI2 Breaker is Closed'
    sheet['A22'].value = 'MSB2-B13 DP2 Breaker is Closed'
    sheet['A23'].value = 'MSB2-B14 DP1 Temp Feed Breaker is racked out'
    sheet['A24'].value = 'Eaton Xpert meter Events light off'
    sheet['A25'].value = 'MSB2-B01 Main Breaker is Closed'
    sheet['A26'].value = 'MSB2-B08 SPD All 3 Green lights (Protected)'
    sheet['A27'].value = 'RPTCS (S1 lights are on)'
    sheet['A28'].value = 'Mode Switch is in the `Closed Transition position'
    sheet['A29'].value = 'Tower Yard' # Rooms
    sheet['A30'].value = 'Check the Towers for any Leaks and noises (Bearings, Belts ect.)'
    sheet['A30'].font = Font(size=7, b=False, i=True, color='FF0000')
    sheet['A30'].alignment = center
    # sheet['A30'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet['A31'].value = 'Tower 1'
    sheet['A32'].value = 'Tower 2'
    sheet['A33'].value = 'Tower 3'
    sheet['A34'].value = 'Tower 4'
    sheet['A35'].value = 'Tower 5'
    sheet['A36'].value = 'Tower 6'
    sheet['A37'].value = 'MUA 3 Status and Coils are clean and OK'
    sheet['A38'].value = 'Fuel Storage Room 1' # Rooms
    sheet['A39'].value = 'Total Gallons of Fuel'
    sheet['A40'].value = 'Leak detection panel'
    sheet['A41'].value = 'Room Temperature'
    sheet['A42'].value = 'EF'
    sheet['A43'].value = '1000 Gallon Day Tank Level'
    sheet['A44'].value = 'Alarms'
    sheet['A45'].value = 'Notes'
    sheet['A45'].alignment = leftTop
    sheet['A45'].font = Font(b=True) 

def pg08_engineer_values():
    columns = range(2, 10, 1)
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    rowsYes = [3, 12, 13, 14, 15, 16, 17, 18, 19, 24, 26, 27, 28, 37, 44]
    rowsCheck = [4, 5, 6, 7, 8, 9, 10, 11, 21, 22, 23, 25, 31, 32, 33, 34, 35, 36, 40, 42, ]
    rowsTG = [39]
    rowsT = [41]
    rowsPer = [43]
    rowsRun = [30]
    rowsOkay = [30]
    '''
    rowsFault = []
    rowsBold = []
    '''
    # Yes or No values
    for col in columns:
        for row in rowsYes:
            sheet.cell(row=row, column=col).value = 'Yes  /  No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 9, color='696969')
    # ✓ / X values
    for col in columns:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).value = '✓  /  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Temperature
    for col in columns:
        for row in rowsT:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '°F'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # TG
    for col in columns:
        for row in rowsTG:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'Gals'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Per
    for col in columns:
        for row in rowsPer:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '% '
            sheet.cell(row=row, column=col).font = Font(size=9, color='696969')
    # Running
    for col in columnEven:
        for row in rowsRun:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = 'Running'
            sheet.cell(row=row, column=col).font = Font(size=8, color='000000')
    # Okay
    for col in columnOdd:
        for row in rowsOkay:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = 'Okay'
            sheet.cell(row=row, column=col).font = Font(size=8, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg08_colored_cells():
    columnsColor = range(1, 10, 1)
    rowsDkGrey = [1, 29, 38]
    for col in columnsColor:
        for row in rowsDkGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx') 

    rowsLtGrey = [2, 20, 30]
    for col in columnsColor:
        for row in rowsLtGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Borders
    rows = range(1, 46)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')
