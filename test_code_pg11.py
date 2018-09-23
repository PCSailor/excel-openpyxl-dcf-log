'''
import os
import openpyxl

# from openpyxl.workbook import Workbook # Todo: Delete once all sheets built
from openpyxl.styles import PatternFill, Protection, Font, GradientFill, NamedStyle, Color, colors
'''

print('\nStart next file, \'test_code_pg11.py\'')
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
wb = load_workbook('Plymouth_Daily_Rounds.xlsx')
# sheet = wb.active
print('Active sheet is', wb.active) # 'Plymouth_Daily_Rounds' active
print('Worksheet list:',  wb.sheetnames) # 'Plymouth_Daily_Rounds', 'Page_11'
indexNumber = wb.worksheets.index(wb['Plymouth_Daily_Rounds'])
print('\'Plymouth_Daily_Rounds\' index number = ', indexNumber) # = 0
indexNumber = wb.worksheets.index(wb['Page_11'])
print('\'Page_11\' index number = ', indexNumber) # 1
'''
Error: This code didn't work to globally change the sheet
# sheet = wb['Page_11'] # Help: why won't this switch the active sheet?
# sheet = wb["Page_11"]
# print('Active sheet is', wb.active) # 'Plymouth_Daily_Rounds' active # Help: need to make 'pg_11' active here
'''
wb.save('Plymouth_Daily_Rounds.xlsx')

# 11-02 is next

def pg11_start():
    print('\'def pg11_start()\' function Sheet list =', wb.sheetnames) # 'Plymouth_Daily_Rounds', 'Page_11'
    print('\n\'def pg11_start()\' function Active sheet #1 =', wb.active) # 'Plymouth_Daily_Rounds' active
    sheet = wb["Page_11"] # Note: This switches active sheet
    print('\'def pg11_start()\' function Active sheet #2 =', sheet) #  # 'Page_11' active
    wb["Page_11"]['A2'].value = 'Note: When doing rounds be aware for unusual smells, sounds, sights, or anything not normal.' # Question: 'wb["Page_11"]' writes to page_11 but 'wb.active' does not, Why?
    wb.save('Plymouth_Daily_Rounds.xlsx')

# back over to 'page_00_index.py', inside the 'if __name__ == "__main__":' code, #2 prints

def pg11_headers():
    # Local Variables
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='bottom')
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), 
                        right=Side(style='thick'), 
                        top=Side(style='thick'), 
                        bottom=Side(style='thick'))
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    # Print Options
    sheet.print_area = 'A1:I43'
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.75
    sheet.page_margins.bottom = 0.75
    sheet.page_margins.header = 0.3
    sheet.page_margins.footer = 0.3
    # Headers & Footers
    sheet.oddHeader.center.text = "&[Tab]"
    sheet.oddHeader.center.size = 24
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "Page &[Page] of &N"
    sheet.oddFooter.left.size = 12
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 12
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000" # 
    print('\n\'def pg11_headers()\' function Active sheet #2 =', sheet)
    print('\'def pg11_headers()\' function Sheet list =', wb.sheetnames)
    wb.save('Plymouth_Daily_Rounds.xlsx')
'''
print('\n11-01-3 (after \'def pg11_headers()\' function is complete)')
print('Active sheet is', sheet)
print('Worksheet list:',  wb.sheetnames)
wb.save('Plymouth_Daily_Rounds.xlsx')
'''
def pg11_merge():
    # Merges 9 cells into 1 in 1 row
    for row in (1, 4, 24, 41, 42, 43):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    # merge 2 cells into 1 in 1 row
    columns = [(col, col+1) for col in range(2, 9, 2)]
    for row in [2, 3, 5, 8, 10, 11, 12, 13, 17, 18, 23, 25, 29, 30, 31, 32, 40]:
        for col1, col2 in columns:
            sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    # merge 2 cells into 1 and 4 cells into 1 cell, all in 1 row
    for row in (22, 39):
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
    # Column width and Row height
    sheet.column_dimensions['A'].width = 30.00
    for col in ['B', 'D', 'F', 'H']:
        sheet.column_dimensions[col].width = 4.00
    for col in ['C', 'E', 'G', 'I']:
        sheet.column_dimensions[col].width = 10.00
    rows = range(1, 43)
    for row in rows:
        sheet.row_dimensions[row].Height = 15.00
    print('\n\'def pg11_merge()\' function Active sheet #2 =', sheet)
    print('\'def pg11_merge()\' function Sheet list =', wb.sheetnames)
    wb.save('Plymouth_Daily_Rounds.xlsx')
'''
print('\n11-01-4 (after \'def pg11_merge()\' function is complete)')
print('Active sheet is', sheet)
print('Worksheet list:',  wb.sheetnames)
wb.save('Plymouth_Daily_Rounds.xlsx')
'''
'''
# Set Named Styles (mutable & used when need to apply formatting to different cells at once)
headerrows = NamedStyle(name='headerrows')
headerrows.font = Font(bold=True, underline='none', sz=12)
headerrows.alignment = center

rooms = NamedStyle(name='rooms')
rooms.font = Font(b=True, sz=12)

rightAlign = NamedStyle(name='rightAlign')
rightAlign.font = Font(b=True, i=True, sz=10)
rightAlign.alignment = Alignment(horizontal='right', vertical='center')
''
# A1
a1 = sheet['A1'] # Note: When doing rounds be aware for unusual smells, sounds, sights, or anything not normal
a1.style = 'rooms'
a1.font = Font(size=11, i=True, color='FF0000')
a1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
# Header Rows
sheet['A2'].style = 'headerrows'
sheet['A3'].style = 'headerrows'
sheet['B3'].style = 'headerrows'
sheet['C3'].style = 'headerrows'
sheet['D3'].style = 'headerrows'
sheet['F3'].style = 'headerrows'
sheet['H3'].style = 'headerrows'
sheet['A2'].style = 'rightAlign'
sheet['A3'].style = 'rightAlign'
# Room Divisions
sheet['A4'].style = 'rooms'
sheet['E4'].style = 'rooms'
sheet['A24'].style = 'rooms'
sheet['A41'].style = 'rooms'

# Set Borders
rows = range(1, 44)
columns = range(1, 10)
for row in rows:
    for col in columns:
        sheet.cell(row, col).border = thin_border
wb.save('Plymouth_Daily_Rounds.xlsx')

# Cell values
a1.value = 'Note: When doing rounds be aware for unusual smells, sounds, sights, or anything not normal.'
sheet['A2'].value = 'Engineer Initials:  '
sheet['A3'].value = 'Time off Round:  '
sheet['B3'].value = '02:00'
sheet['D3'].value = '08:00'
sheet['F3'].value = '14:00'
sheet['H3'].value = '20:00'
# Server Room #1_Equipment
sheet['A4'].value = 'Server Room 1'
sheet['A5'].value = 'Tear off sticky mat (Battery Room)'
sheet['A6'].value = 'CRAC 24'
sheet['A7'].value = 'CRAC 23'
sheet['A8'].value = 'SR1 CHW Loop'
sheet['A9'].value = 'CRAC 04'
sheet['A10'].value = 'PDU 11'
sheet['A11'].value = 'PDU 09'
sheet['A12'].value = 'PDU 02'
sheet['A13'].value = 'PDU 04'
sheet['A14'].value = 'CRAC 26'
sheet['A15'].value = 'CRAC 05'
sheet['A16'].value = 'CRAC 06'
sheet['A17'].value = 'PDU 01'
sheet['A18'].value = 'PDU 08'
sheet['A19'].value = 'CRAC 33'
sheet['A20'].value = 'CRAC 07'
sheet['A21'].value = 'Humidifier'
sheet['A22'].value = 'FM 200 (2 tanks)'
sheet['A23'].value = 'Tear off sticky mat (Hallway)'

# Server Room #3_Equipment
sheet['A24'].value = 'Server Room 3'
sheet['A25'].value = 'Tear of sticky mat (Hallway)'
sheet['A26'].value = 'CRAC 10'
sheet['A27'].value = 'CRAC 22'
sheet['A28'].value = 'CRAC 31'
sheet['A29'].value = 'PDU 23'
sheet['A30'].value = 'PDU 22'
sheet['A31'].value = 'PDU 03'
sheet['A32'].value = 'PDU 10'
sheet['A33'].value = 'CRAC 11'
sheet['A34'].value = 'CRAC 12'
sheet['A35'].value = 'CRAC 13'
sheet['A36'].value = 'CRAC 14'
sheet['A37'].value = 'CRAC 30'
sheet['A38'].value = 'Humidifier'
sheet['A39'].value = 'FM 200'
sheet['A40'].value = 'Tear off sticky mat (Loading Dock)'
sheet['A41'].value = 'Notes:' # StretchGoal: Increase row height, delete comment rows below
sheet['A42'].value = ''
sheet['A43'].value = ''

# Engineering Values
# Yes or No values
columns = [2, 4, 6, 8]
rows = [5, 23, 25, 40]
# cells = []
for col in columns:
    for row in rows:
        sheet.cell(row=row, column=col).value = 'Yes  /  No'
        sheet.cell(row=row, column=col).alignment = center
        sheet.cell(row=row, column=col).font = Font(size = 8, i=True, color='000000')

# ✓ X values
rowsCheck = [6, 7, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39]
for col in columns:
    for row in rowsCheck:
        # print(col, row)
        sheet.cell(row=row, column=col).value = '✓  X'
        sheet.cell(row=row, column=col).alignment = center
        sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')

# RH%
columnodd = [3, 5, 7, 9]
rowsRH = [21, 38]
for col in columnodd:
    for row in rowsRH:
        # print(col, row)
        sheet.cell(row=row, column=col).value = '%RH'
        sheet.cell(row=row, column=col).alignment = right
        sheet.cell(row=row, column=col).font = Font(size=8, color='000000')

# Hz
rowsHZ = [6, 7, 9, 14, 15, 16, 19, 20, 26, 27, 28, 33, 34, 35, 36, 37]
for col in columnodd:
    for row in rowsHZ:
        # print(col, row)
        sheet.cell(row=row, column=col).value = 'Hz'
        sheet.cell(row=row, column=col).alignment = right
        sheet.cell(row=row, column=col).font = Font(size=8, color='000000')

# D/P
rowsRH = [8]
for col in columnEven:
    for row in rowsRH:
        # print(col, row)
        sheet.cell(row=row, column=col).value = 'D/P'
        sheet.cell(row=row, column=col).alignment = right
        sheet.cell(row=row, column=col).font = Font(size=8, color='000000')

# Colored Cells
rowsColor = [1, 2, 3, 4, 24, 41]
columnsColor = [1, 2, 4, 6, 8]
for col in columnsColor:
    for row in rowsColor:
        # print(col, row)
        sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
'''

print('\n11-02')
'''
wb = load_workbook('Plymouth_Daily_Rounds.xlsx')

print('\'Page 11\' sheet dimensions =', wb.active.dimensions)

print('End pg11 worksheet list:',  wb.sheetnames)
'''
# sheet = wb["Page_11"]
print('End pg11 active sheet is', wb.active) # 'Plymouth_Daily_Rounds' active
print('End file, \'test_code_pg11.py\'', wb.sheetnames) # 'Plymouth_Daily_Rounds', 'Page_11'
# wb.save('Plymouth_Daily_Rounds.xlsx')
