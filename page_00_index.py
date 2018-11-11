#! python3
'''
* Page Notes:
    * Page_01: Complete 
    * Page_02: Complete 
    * Page_03: Complete 
    * Page_04: Complete 
    * Page_05: Complete 
    * Page_06: Complete 
    * Page_07:
    * Page_08:
    * Page_09:
    * Page_10:
    * Page_11: Complete 
* Code Structure Outline
    * SheBang
    * Comments
    * Imports
    * Global Variables
    * Headers
    * Merge
        * Dimensions # Todo: set as own def
        * Page Fonts # Todo: set as own def
    * NamedStyles
    * Cell Values
    * Engineer Values
        * Cell-specific adjustments
    * Colored Cells
    * Borders
END
* The Golden Loop:
                        def PAGE-NAME-HERE_cell_values():
                            # Cell values
                            values = ['', '', '', '', 'Notes']
                            def update_vertically(sheet, col, row, values):
                                row = int(row)
                                for i, value in enumerate(values):
                                    sheet['{}{}'.format(col, row+i)].value = value
                            update_vertically(sheet, 'A', 2, values)

* (Python on Windows)[https://docs.python.org/2/faq/windows.html]
* (openpyxl.worksheet package_Submodules)[https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.html])
* (Openpyxl: Working with Styles)[https://openpyxl.readthedocs.io/en/stable/styles.html?highlight=cell%20border#introduction]
* (RGB Colors)[https://www.rapidtables.com/web/color/RGB_Color.html]

* Setting Borders - The merged cell behaves similar to other cell ojects. Its value and format is defined in its top-left cell. In order to change the border of the whole merged cell, change the border of its top-left cell.
    * Border properties: {'mediumDashDotDot', 'dashDotDot', 'dotted', 'hair', 'slantDashDot', 'mediumDashed', 'thin', 'medium', 'double', 'thick', 'dashDot', 'dashed', 'mediumDashDot'}
* Document Color Selections:
    * Permanent text set to black, #000000
    * Light-Grey text set to 696969 (or DCDCDC)
    * Dark Grey box set to C0C0C0
    * Light Grey box set to DCDCDC

* Left-To-Do: 
    * Clear opening error (see pics)
'''
print('\n\'  page_00_index.py\' is run')
# import prettyprinter
import sys
import os
cwd = os.getcwd()
import openpyxl
from openpyxl.workbook import Workbook # what is 'import load_workbook'?
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle, Color, colors
# from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill, Protection, GradientFill, Color, colors

# Set directory # Todo: Change directory
# Direct Path
# os.chdir('D:\Software\Python\python_excel\python_openpyxl_dcflog_updated')
# absolute path
os.chdir(os.path.abspath(os.path.dirname(__file__)))
'''
os.chdir = change working directory
os.path.abspath = return the absolute path
os.path.dirname = return the name portion of the object
__file__ = The path to where the module data is stored (not set for built-in modules)
'''
print('  CWD =', cwd) 
    # Same results: 
        # print(os.getcwd())
        # print('Current Working Directory is %s:' % cwd)

# create workbook
wb = openpyxl.Workbook()
print('  Workbook Type =', type(wb))
print('  Worksheet list-before:', wb.sheetnames)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Set sheets
sheet = wb.active
sheet.title = 'Page_01'
# Create Sheet
sheet = wb.create_sheet(title='Page_02', index=1)
sheet = wb.create_sheet(title='Page_03', index=2)
sheet = wb.create_sheet(title='Page_04', index=3)
sheet = wb.create_sheet(title='Page_05', index=4)
sheet = wb.create_sheet(title='Page_06', index=5)
sheet = wb.create_sheet(title='Page_07', index=6)
sheet = wb.create_sheet(title='Page_08', index=7)
sheet = wb.create_sheet(title='Page_09', index=8)
sheet = wb.create_sheet(title='Page_10', index=9)
sheet = wb.create_sheet(title='Page_11', index=10)
# sheet = wb.create_sheet(title='Loops', index=11)
# sheet = wb.create_sheet(title='test_code', index=11)
print('  Worksheet list-after:',  wb.sheetnames) # 
wb.save('Plymouth_Daily_Rounds.xlsx')

if __name__ == "__main__":
    from page_01_bms_commandctr import pg01_headers, pg01_merge, pg01_namedstyle, pg01_cell_values, pg01_engineer_values, pg01_colored_cells
    pg01_headers()
    pg01_merge()
    pg01_namedstyle()
    pg01_cell_values()
    pg01_engineer_values()
    pg01_colored_cells()
    # 
    from page_02_server_2_mdf_estbattrm import pg02_headers, pg02_merge, pg02_namedstyle, pg02_cell_values, pg02_engineer_values, pg02_colored_cells
    # pg02_start()
    pg02_headers()
    pg02_merge()
    pg02_namedstyle()
    pg02_cell_values()
    pg02_engineer_values()
    pg02_colored_cells()
    #
    from page_03_eastelectrical import pg03_headers, pg03_merge, pg03_namedstyle, pg03_cell_values, pg03_engineer_values, pg03_colored_cells
    pg03_headers()
    pg03_merge()
    pg03_namedstyle()
    pg03_cell_values()
    pg03_engineer_values()
    pg03_colored_cells()
    # 
    from page_04_firepprm_docking import pg04_headers, pg04_merge, pg04_namedstyle, pg04_cell_values, pg04_colored_cells, pg04_engineer_values
    pg04_headers()
    pg04_merge()
    pg04_namedstyle()
    pg04_cell_values()
    pg04_engineer_values()
    pg04_colored_cells()
    # 
    from page_05_machinery_01 import pg05_headers, pg05_merge, pg05_namedstyle, pg05_cell_values, pg05_engineer_values, pg05_colored_cells
    pg05_headers()
    pg05_merge()
    pg05_namedstyle()
    pg05_cell_values()
    pg05_engineer_values()
    pg05_colored_cells()
    #  
    from page_06_machinery_02 import pg06_headers, pg06_merge, pg06_cell_values, pg06_engineer_values, pg06_colored_cells, pg06_namedstyle
    pg06_headers()
    pg06_merge()
    pg06_namedstyle()
    pg06_cell_values()
    pg06_engineer_values()
    pg06_colored_cells()
    # 
    from page_07_machinery_03 import pg07_headers, pg07_merge, pg07_namedstyle, pg07_cell_values, pg07_engineer_values, pg07_colored_cells
    pg07_headers()
    pg07_merge()
    pg07_namedstyle()
    pg07_cell_values()
    pg07_engineer_values()
    pg07_colored_cells()
    #
    from page_08_toweryard import pg08_headers, pg08_merge, pg08_namedstyle, pg08_cell_values, pg08_engineer_values, pg08_colored_cells
    pg08_headers()
    pg08_merge()
    pg08_namedstyle()
    pg08_cell_values()
    pg08_engineer_values()
    pg08_colored_cells()
    #
    from page_09_generators import pg09_headers, pg09_merge, pg09_namedstyle, pg09_cell_values, pg09_engineer_values, pg09_colored_cells
    pg09_headers()
    pg09_merge()
    pg09_namedstyle()
    pg09_cell_values()
    pg09_engineer_values()
    pg09_colored_cells()
    #
    from page_10_westelectrical_westbattery import pg10_headers, pg10_merge, pg10_namedstyle, pg10_cell_values, pg10_engineer_values, pg10_colored_cells
    pg10_headers()
    pg10_merge()
    pg10_namedstyle()
    pg10_cell_values()
    pg10_engineer_values()
    pg10_colored_cells()#
    # breakpoint()
    from page_11_server_1_3 import pg11_headers, pg11_merge, pg11_namedstyle, pg11_cell_values, pg11_engineer_values, pg11_colored_cells
    pg11_headers()
    pg11_merge()
    pg11_namedstyle()
    pg11_cell_values()
    pg11_engineer_values()
    pg11_colored_cells()
    '''
    #
    from page_12_openpyxl_code_dump import pg12_headers, pg12_merge, pg12_namedstyle, pg12_cell_values, pg12_engineer_values, pg12_colored_cells
    pg12_headers()
    pg12_merge()
    pg12_namedstyle()
    pg12_cell_values()
    pg12_engineer_values()
    pg12_colored_cells()
    #
    from Loops import loops_cell_values, pg09_cell_values
    loops_cell_values()
    pg09_cell_values()
    # 
    from test_code import pg_tc_headers, pg_tc_merge, pg_tc_namedstyle, pg_tc_cell_values, pg_tc_engineer_values, pg_tc_colored_cells
    pg_tc_headers()
    pg_tc_merge()
    pg_tc_namedstyle()
    pg_tc_cell_values()
    pg_tc_engineer_values()
    pg_tc_colored_cells()
    * With 'page_11_server_1_3' last in 'if __name__ == "__main__":', this page has errors in the borders of merged cells
    ** With this last in 'if __name__ == "__main__":' the page formats perfectly in MS Excel
    '''
print('\nWorksheet list at end of \'page_00\':', wb.sheetnames) # 'Plymouth_Daily_Rounds', 'Page_11'