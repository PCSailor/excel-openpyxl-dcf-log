#! python3
'''
* Server Room 2
* MDF Room
* Fire Pump Room
'''
print('\nStart next file, \'page_02\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_02"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
columns = [(col, col+1) for col in range(2, 9, 2)]
center = Alignment(horizontal='center', vertical='center')
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# Local Variable
values = ['Mechanical / Chill Water Units Room'

]

def pg02_headers():
    # Print Options
    sheet.print_area = 'A1:I46' # TODO: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000" # 
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg02_merge():
    # Merges 9 cells into 1 in 1 row
    for row in (1, 30, 38, 39, 47, 48, 49):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    # merge 2 cells into 1 in 1 row
    for row in [6, 7, 8, 9, 13, 18, 19, 23, 25,26, 27, 28, 29, 31, 32, 36, 40, 41, 42, 43, 45, 46]:
        for col1, col2 in columns:
            sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    # merge 2 cells into 1 and 4 cells into 1 cell, all in 1 row
    for row in (10, 21, 35):
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)

    # Dimensions_Columns & Rows
    sheet.column_dimensions['A'].width = 36.00
    for col in ['B', 'D', 'F', 'H']:
        sheet.column_dimensions[col].width = 4.25
    for col in ['C', 'E', 'G', 'I']:
        sheet.column_dimensions[col].width = 11.00
    rows = range(1, 46)
    for row in rows:
        sheet.row_dimensions[row].height = 15.00
    sheet.row_dimensions[44].height = 30.00
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Page Font
    rows = range(1, 50)
    columnFont = range(1, 10)
    for row in rows:
        for col in columnFont:
            sheet.cell(row, col).font = Font(size = 10, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg02_namedstyle():
    # Set Named Styles (mutable & used when need to apply formatting to different cells at once)
    # Room Divisions
    sheet['A1'].style = 'rooms'
    sheet['A30'].style = 'rooms'
    sheet['A38'].style = 'rooms'
    sheet['A39'].style = 'subtitles'
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg02_cell_values():
    # Cell values
    ''' def populate(sheet, col, row, values):
            row = int(row)
            for i, value in enumerate(values):
                    sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list '''
    #
    sheet['A1'].value = 'Server Room 2'
    sheet['A2'].value = 'CRAC 29'
    sheet['A3'].value = 'CRAC 21'
    sheet['A4'].value = 'CRAC 32'
    sheet['A5'].value = 'Humidifier'
    sheet['A6'].value = 'PDU 21'
    sheet['A7'].value = 'PDU 20'
    sheet['A8'].value = 'PDU 06'
    sheet['A9'].value = 'PDU 14'
    sheet['A10'].value = 'FM 200'
    sheet['A11'].value = 'CRAC 27'
    sheet['A12'].value = 'CRAC 28'
    sheet['A13'].value = 'SR2 CHW Loop'
    sheet['A14'].value = 'CRAC 17'
    sheet['A15'].value = 'CRAC 16'
    sheet['A16'].value = 'CRAC 19'
    sheet['A17'].value = 'CRAC 18'
    sheet['A18'].value = 'PDU 17'
    sheet['A19'].value = 'PDU 16'
    sheet['A20'].value = 'CRAC 34'
    sheet['A21'].value = 'FM 200'
    sheet['A22'].value = 'CRAC 15'
    sheet['A23'].value = 'CRAC 25'
    sheet['A24'].value = 'CRAC 20'
    sheet['A25'].value = 'PDU 19'
    sheet['A26'].value = 'PDU 18'
    sheet['A27'].value = 'PDU 07'
    sheet['A28'].value = 'PDU 15'
    sheet['A29'].value = 'Tear off Sticky Mat'
    sheet['A30'].value = 'MDF'
    sheet['A31'].value = 'Tear off Sticky Mat'
    sheet['A32'].value = 'PDU 12'
    sheet['A33'].value = 'CRAC 08'
    sheet['A34'].value = 'Humidifier'
    sheet['A35'].value = 'FM 200'
    sheet['A36'].value = 'PDU 05'
    sheet['A37'].value = 'CRAC 09'
    sheet['A38'].value = 'East Battery Room'
    sheet['A39'].value = 'Rail 2 Batteries'
    sheet['A40'].value = 'CU2 Battery Circuit Breaker is closed'
    sheet['A41'].value = 'Eagle Eye Computer Alarms'
    sheet['A42'].value = ''
    sheet['A43'].value = ''
    sheet['A44'].value = 'DC Ground Fault Module below 6MA\n(Pre-alarm = 10MA, Alarm = 20MA)'
    sheet['A45'].value = 'Spare Battery Charger'
    # sheet['A46'].value = '' # merge
    ''' sheet['A47'].value = 'Notes:'
    sheet['A47'].alignment = leftTop
    sheet['A47'].font = Font(b=True) '''
    wb.save('Plymouth_Daily_Rounds.xlsx')

    ''' sheet['D10'].value = '✓ X'
    sheet['D21'].value = '✓ X'
    sheet['D35'].value = '✓ X'
    sheet['D10'].font = Font(size=8, color='DCDCDC')
    sheet['D21'].font = Font(size=8, color='DCDCDC')
    sheet['D35'].font = Font(size=8, color='DCDCDC')
    sheet['D10'].alignment = center
    sheet['D21'].alignment = center
    sheet['D35'].alignment = center '''

def pg02_engineer_values():
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    rowsYes = [29, 31]
    rowsRH = [5, 34]
    rowsCheck = [2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 32, 33, 34, 35, 36, 37, 40, 44]
    rowsHz = [2, 3, 4, 11, 12, 14, 15, 16, 17, 20, 22, 23, 24, 33, 37]
    rowsVac = [41, 45]
    rowsR = [42]
    rowsT = [43]
    rowsI = [46]
    rowsPsi = [13]
    rowsFM = [10, 21, 35]
    columnFM = [2, 6]
    # Yes or No values
    for col in columnEven:
        for row in rowsYes:
            sheet.cell(row=row, column=col).value = 'Yes  /  No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 9, color='696969')
    # ✓ X values
    for col in columnEven:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).value = '✓   X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # RH%
    for col in columnOdd:
        for row in rowsRH:
            sheet.cell(row=row, column=col).value = '%RH'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Hz
    for col in columnOdd:
        for row in rowsHz:
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Voltage
    for col in columnEven:
        for row in rowsVac:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'vDC'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Resistance
    for col in columnEven:
        for row in rowsR:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'Ω'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Temperature
    for col in columnEven:
        for row in rowsT:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '°F'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Amps
    for col in columnEven:
        for row in rowsI:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'amps'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # PSI
    for col in columnEven:
        for row in rowsPsi:
            sheet.cell(row=row, column=col).value = 'PSI'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
     # FM200
    for col in columnFM:
        for row in rowsFM:
            sheet.cell(row=row, column=col).value = ''
    # Cell-specific adjustments
    sheet.merge_cells(start_row=41, start_column=1, end_row=43, end_column=1)
    sheet.merge_cells(start_row=45, start_column=1, end_row=46, end_column=1)
    sheet.cell(row=41, column=1).alignment = center
    sheet.cell(row=45, column=1).alignment = center
    sheet.cell(row=47, column=1).alignment = leftTop
    sheet['A44'].alignment = wrap # auto-adjusts row height?

def pg02_colored_cells():
    rowsDkGrey = [1, 30, 38]
    rowsLtGrey = [39]
    columnsColor = [1, 2, 4, 6, 8]
    rowsBlack = [10, 21, 35]
    columnsBlack = [2, 6, 8]
    for col in columnsColor:
        for row in rowsDkGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    for col in columnsColor:
        for row in rowsLtGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    for col in columnsBlack:
        for row in rowsBlack:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')
    
    # Borders
    rows = range(1, 48)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')