'''
# Center all cells vertical & horizontal
allCells = sheet['A1':'I43']
for row in allCells:
    for cell in row:
        cell.alignment = al # THIS IS WORKING BUT CODE BELOW CHANGES THE ALIGNMENT

Highlight: Merge Bug
(Revised method of constructing borders for merged cells)[https://bitbucket.org/openpyxl/openpyxl/pull-requests/274/revised-method-of-constructing-borders-for/diff]
(Styling Merged Cells isn't working)[https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working]
'''
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# Read an existing workbook ()[https://openpyxl.readthedocs.io/en/2.4/usage.html?highlight=load_workbook}]
from openpyxl import load_workbook
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)

sheet = wb.create_sheet('Colors', 1)

# sheet.title = 'Colors'
# sheet = [1]
print('all.sheetnames:', wb.sheetnames)
print(sheet, ' is active')

# Sheet: Colors
# Styling
# Samples

# Darkest to lightest
sheet['b3'].value = '696969'
sheet['b2'].value = '808080'
sheet['b4'].value = 'A9A9A9'
sheet['b5'].value = 'C0C0C0'
# these two look beige
sheet['b1'].value = 'D3D3D3'
sheet['b6'].value = 'DCDCDC'

sheet['a1'].fill = PatternFill(fgColor='D3D3D3', fill_type = 'solid')
sheet['a2'].fill = PatternFill(fgColor='808080', fill_type = 'solid')
sheet['a3'].fill = PatternFill(fgColor='696969', fill_type = 'solid')
sheet['a4'].fill = PatternFill(fgColor='A9A9A9', fill_type = 'solid')
sheet['a5'].fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
sheet['a6'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
print('\'openpyxl_code_dump.py\' run complete with sheet dimensions of ', sheet.dimensions)
wb.save('server_1_and_3.xlsx')