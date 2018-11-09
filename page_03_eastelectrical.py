#! python3
'''
* East Battery Room
* East UPS Room

* Bug: sheet['A4/5/6'] merge leave the left border on rows 5 & 6 blank, no border
* Bug: sheet['A18/19/20'] merge leave the left border on rows 19 & 20 blank, no border
'''
print('\nStart next file, \'page_03\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_03"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
center = Alignment(horizontal='center', vertical='center')
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# Local Variable
values = ['Mechanical / Chill Water Units Room'

]

def pg03_headers():
    # Print Options
    sheet.print_area = 'A1:I43' # TODO: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg03_merge():
    # Merges 9 cells into 1 in 1 row
    for row in (1, 8, 10, 11, 22, 28, 40, 44):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    # merge 2 cells into 1 in 1 row
    columns = [(col, col+1) for col in range(2, 9, 2)]
    for row in [2, 3, 4, 5, 6, 7, 9, 12, 13, 14, 15, 18, 19, 20, 21, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 41, 42, 43]:
        for col1, col2 in columns:
            sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    sheet.merge_cells(start_row=4, start_column=1, end_row=6, end_column=1)
    sheet.merge_cells(start_row=18, start_column=1, end_row=20, end_column=1)
    
    # Dimensions_Columns & Rows
    sheet.column_dimensions['A'].width = 37.75
    for col in ['B', 'D', 'F', 'H']:
        sheet.column_dimensions[col].width = 5.0
    for col in ['C', 'E', 'G', 'I']:
        sheet.column_dimensions[col].width = 10.0
    rows = range(1, 43)
    for row in rows:
        sheet.row_dimensions[row].height = 15.00

    sheet['A7'].alignment = left # Note: var 'left' includes 'wrap' but wrap does not auto-adjusts row height
    sheet['A7'].alignment = left
    sheet['A23'].alignment = left
    sheet['A26'].alignment = left
    sheet['A27'].alignment = left
    sheet['A29'].alignment = left
    sheet['A38'].alignment = left
    sheet.row_dimensions[7].height = 30.00
    sheet.row_dimensions[23].height = 30.00
    sheet.row_dimensions[26].height = 30.00
    sheet.row_dimensions[27].height = 30.00
    sheet.row_dimensions[29].height = 30.00
    sheet.row_dimensions[38].height = 30.00
    sheet.row_dimensions[44].height = 30.00

    
    # Cell-specific adjustments
    # sheet.cell(row=4, column=1).alignment = center # A4
    # sheet.cell(row=18, column=1).alignment = center # A18
    wb.save('Plymouth_Daily_Rounds.xlsx')

    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Page Font
    rows = range(1, 50)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).font = Font(size = 10, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg03_namedstyle():
    # Styles
    sheet['A1'].style = 'rooms'
    sheet['A8'].style = 'rooms'
    sheet['B21'].style = 'rightAlign' # Todo: Add into forLoop
    sheet['B24'].style = 'rightAlign'
    sheet['B25'].style = 'rightAlign'
    sheet['B27'].style = 'rightAlign'
    sheet['A10'].style = 'subtitles'
    sheet['A22'].style = 'subtitles'
    sheet['A28'].style = 'subtitles'
    sheet['A40'].style = 'subtitles'
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg03_cell_values():
    # Cell values
    ''' def populate(sheet, col, row, values):
            row = int(row)
            for i, value in enumerate(values):
                    sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list '''
    #
    sheet['A1'].value = 'Rail 3 Batteries'
    sheet['A2'].value = 'Room Temperature'
    sheet['A3'].value = 'CU3 Battery Breaker is Closed'
    sheet['A4'].value = 'Eagle Eye Computer Alarms' # Todo: Merge A4-A6
    # sheet['A5'].value = ''
    # sheet['A6'].value = ''
    sheet['A7'].value = 'DC ╧ Fault Module below 6MA\n(Pre-alarm = 10MA, Alarm = 20MA)' # alt 207
    sheet['A8'].value = 'East UPS Room'
    sheet['A9'].value = 'Battery Room Hydrogen Monitor % Levels'
    sheet['A10'].value = 'Rail 2'
    sheet['A11'].value = '** Ensure key is in locked position before touching STS display screen!! **'
    sheet['A11'].font = Font(size=12, b=True, i=True, color='FF0000')
    sheet['A11'].alignment = center
    sheet['A12'].value = 'STS-2A on preferred Source #1'
    sheet['A13'].value = 'STS-2B on preferred Source #1'
    sheet['A14'].value = 'STS-2C on preferred Source #1'
    sheet['A15'].value = 'STS-2D on preferred Source #1'
    sheet['A16'].value = 'CRAC 37'
    sheet['A17'].value = 'CRAC 36'
    sheet['A18'].value = 'CU2-M1_UPS 2' # Todo: Merge A18-A20
    # sheet['A19'].value = ''
    # sheet['A20'].value = ''
    sheet['A21'].value = 'MBB Module Battery Breaker is Closed'
    sheet['A22'].value = 'CI2'
    sheet['A23'].value = 'CI2-B08 Breaker SPD with 3 Green lights (Protected)'
    sheet['A24'].value = 'CI2-B11 CU2-M1 Input Breaker is Closed'
    sheet['A25'].value = 'CI2-B06 Static Bypass Breaker is Closed'
    sheet['A26'].value = 'Eaton Xpert Meter Events light is OFF (User=X & PW=X)'
    sheet['A27'].value = 'CI2-B05 Maintenance Bypass Breaker is Closed'
    sheet['A28'].value = 'CO2'
    sheet['A29'].value = 'Eaton Breaker Interface Module II Alarm light is OFF'
    sheet['A30'].value = 'CO2-B18 Spare is Open'
    sheet['A31'].value = 'CO2-B11 STS-2A is Closed'
    sheet['A32'].value = 'CO2-B12 STS-2B is Closed'
    sheet['A33'].value = 'CO2-B17 Spare is Open'
    sheet['A34'].value = 'CO2-B13 STS-1A is Closed'
    sheet['A35'].value = 'CO2-B14 STS-3B is Closed'
    sheet['A36'].value = 'CO2-B15 STS-2C is Closed'
    sheet['A37'].value = 'CO2-B16 STS-2D is Closed'
    sheet['A38'].value = 'Eaton Xpert Meter Events light is OFF (User=X & PW=X)'
    sheet['A39'].value = 'CO2-B01 CC-CO Isolation switch is Closed'
    sheet['A40'].value = 'CC3' # 
    sheet['A41'].value = 'CC3-B05 (MBB) Breaker is Open'
    sheet['A42'].value = 'CC3-B01 (MIB) Breaker is Closed'
    sheet['A43'].value = 'CC3-B99 (LBB) Breaker is Open'

    ''' sheet['A44'].value = 'Notes:'
    sheet['A44'].alignment = leftTop
    sheet['A44'].font = Font(b=True)
    sheet['A44'].alignment = leftTop # A44 '''

    ''' FixMe: Delete this??
    sheet['B16'].value = '✓ X'
    sheet['B16'].font = Font(size=8, color='DCDCDC')
    sheet['B17'].value = '✓ X'
    sheet['B17'].font = Font(size=8, color='DCDCDC') '''
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg03_engineer_values():
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    rowYes = [7, 23, 26, 29, 38]
    rowCheck = [3, 12, 13, 14, 15, 16, 17, 21, 24, 25, 27, 30, 31, 32, 33, 34, 35, 36, 37, 39, 41, 42, 43]
    rowHZ = [16, 17]
    rowVac = [4]
    rowR = [5]
    rowT = [2, 6]
    rowEf = [9]
    rowkW = [18]
    rowkVA = [19]
    rowsAct = [20]
    # Yes or No values
    for col in columnEven:
        for row in rowYes:
            sheet.cell(row=row, column=col).value = 'Yes  /  No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 9, color='696969')
    # ✓ X values
    for col in columnEven:
        for row in rowCheck:
            sheet.cell(row=row, column=col).value = '✓   X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Hz
    for col in columnOdd:
        for row in rowHZ:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Voltage
    for col in columnEven:
        for row in rowVac:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'vDC'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Resistance
    for col in columnEven:
        for row in rowR:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'Ω'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Temperature
    for col in columnEven:
        for row in rowT:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '°F'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # EF%
    for col in columnEven:
        for row in rowEf:
            sheet.cell(row=row, column=col).alignment = ctrdwn
            sheet.cell(row=row, column=col).value = 'EF4%    /    EF5%'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # kW-Out
    for col in columnEven:
        for row in rowkW:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'kW-Out'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # kVA-Out
    for col in columnEven:
        for row in rowkVA:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'kVA-Out'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Active Events
    for col in columnEven:
        for row in rowsAct:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = 'Active Events? Y/N'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')

def pg03_colored_cells():
    rowsDkGrey = [8]
    columnsColor = [1, 2, 3, 4, 5]
    for col in columnsColor:
        for row in rowsDkGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    rowsLtGrey = [1, 10, 22, 28, 40]
    for col in columnsColor:
        for row in rowsLtGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')
    
    # Borders
    rows = range(1, 45)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')