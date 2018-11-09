#!/usr/bin/env python3
'''
* 
wb.save('Plymouth_Daily_Rounds.xlsx')
'''
print('\nStart next file, \'page_07\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_07"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
columns = [(col, col+1) for col in range(2, 6, 2)]
center = Alignment(horizontal='center', vertical='center')
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# Local Variable
'''
values = ['Mechanical / Chill Water Units Room', 'Lakos filter (Big) PSI'
    sheet['A3'].value = 'Lakos filter (Small) PSI'
    sheet['A4'].value = 'Flow Max filter '
    sheet['A5'].value = 'West side CHW'
    sheet['A5'].alignment = left
    sheet['A6'].value = 'Refrigerant Monitor'
    # sheet['A7'].value = 'Refrigerant Monitor'
    sheet['A8'].value = 'HP LL- 1  Ok  (Fan is ok, pipe sweating noticed, HP operation)'
    sheet['A9'].value = 'CHW Temp West'
    sheet['A10'].value = 'Rail 1 UPS Annunciator panel'
    sheet['A11'].value = 'Rail 2 UPS Annunciator panel'
    sheet['A12'].value = 'Rail 3 UPS Annunciator panel'
    sheet['A13'].value = 'DP 2 Breakers' # Subtitles
    sheet['A14'].value = 'DP2-B17 Spare Breaker is Open'
    sheet['A15'].value = 'DP2-B18 ATS-HPB1-C Breaker is Closed'
    sheet['A16'].value = 'DP2-B19 ATS-HPA1-C Breaker is Closed'
    sheet['A17'].value = 'DP2-B20 Spare Breaker is Open'
    sheet['A18'].value = 'DP2-B08 SPD All 3 Green lights (Protected)'
    sheet['A19'].value = 'DP2-B11 ATS-HPB-H Breaker is Closed'
    sheet['A20'].value = 'DP2-B12 CHILLER 2 Breaker is Closed'
    sheet['A21'].value = 'DP2-B13 ATS-HPB-C Breaker is Closed'
    sheet['A22'].value = 'DP2-B14 SPARE Breaker is Open'
    sheet['A23'].value = 'DP2-B15 T-LBP-0 Breaker is Closed'
    sheet['A24'].value = 'DP 1 Breakers' # Subtitles
    sheet['A25'].value = 'DP1-B21 Alt. Feed from MSB3 Breaker is Open'
    sheet['A26'].value = 'DP1-B16 T-PPLL01 Breaker is Closed'
    sheet['A27'].value = 'DP1-B17 ATS-HPD-H Breaker is Closed'
    sheet['A28'].value = 'DP1-B18 ATS-HPC1-C Breaker is Closed'
    sheet['A29'].value = 'DP1-B19 ATS-HPA1-C Breaker is Closed'
    sheet['A30'].value = 'DP1-B20 SPARE Breaker is Open'
    sheet['A31'].value = 'DP1-B08 SPD All 3 Green lights (Protected)'
    sheet['A32'].value = 'DP1-B11 ATS-HPB-H Breaker is Closed'
    sheet['A33'].value = 'DP1-B12 CHILLER 1 Breaker is Closed'
    sheet['A34'].value = 'DP1-B13 ATS-HPA-C Breaker is Closed'
    sheet['A35'].value = 'DP1-B14 ATS-HPA-H Breaker is Closed'
    sheet['A36'].value = 'DP1-B15 T-HLA-0 Breaker is Closed'
    sheet['A37'].value = 'MSB 3' # Subtitles
    sheet['A38'].value = 'MSB3-B12 CI3 Breaker is Closed'
    sheet['A39'].value = 'MSB3-B13 DP3 Breaker is Closed'
    sheet['A40'].value = 'Eaton Xpert meter Events light OFF  (User is X and PW is X)'
    sheet['A41'].value = 'MSB3-B01 Main Breaker is Closed'
    sheet['A42'].value = 'MSB3-B08 SPD All 3 Green lights (Protected)'
    sheet['A43'].value = 'RPTCS (S1 lights are on)'
    sheet['A44'].value = 'Mode Switch is in the Closed Transition position'
    sheet['A45'].value = 'Notes'

]
'''
def pg07_headers():
    # Print Options
    sheet.print_area = 'A1:E45' # Todo: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg07_merge():
    columns = [(col, col+1) for col in range(2, 6, 2)]
    # merge 2 cells in 1 column
    # sheet.merge_cells(start_row=6, start_column=1, end_row=7, end_column=1)
    # Merges 5 cells into 1 in 1 row
    for row in [1, 13, 24, 37, 45]:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    # merge 2 cells into 1 in 1 row
    for row in []:
            for col1, col2 in columns:
                sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Dimensions_Columns & Rows
    sheet.column_dimensions['A'].width = 50.00
    for col in ['B', 'C', 'D', 'E']:
        sheet.column_dimensions[col].width = 9.5
    rows = range(1, 55)
    for row in rows:
        sheet.row_dimensions[row].height = 15.00
    sheet.row_dimensions[1].height = 30.00
    sheet.row_dimensions[46].height = 30.00
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Page Font
    rows = range(1, 55)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).font = Font(size = 10, i=False, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg07_namedstyle():
    # Styles
    sheet['A1'].style = 'rooms'
    sheet['A13'].style = 'subtitles'
    sheet['A24'].style = 'subtitles'
    sheet['A37'].style = 'subtitles'
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg07_cell_values():
    # Cell values
    ''' def populate(sheet, col, row, values):
            row = int(row)
            for i, value in enumerate(values):
                    sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list '''
    #
    sheet['A1'].value = 'Mechanical / Chill Water Units Room continued'
    sheet['A2'].value = 'Lakos filter (Big) PSI'
    sheet['A3'].value = 'Lakos filter (Small) PSI'
    sheet['A4'].value = 'Flow Max filter '
    sheet['A5'].value = 'West side CHW'
    sheet['A5'].alignment = left
    sheet['A6'].value = 'Refrigerant Monitor'
    sheet['A7'].value = 'Refrigerant Monitor'
    sheet['A8'].value = 'HP LL- 1  Ok  (Fan is ok, pipe sweating noticed, HP operation)'
    sheet['A9'].value = 'CHW Temp West'
    sheet['A10'].value = 'Rail 1 UPS Annunciator panel'
    sheet['A11'].value = 'Rail 2 UPS Annunciator panel'
    sheet['A12'].value = 'Rail 3 UPS Annunciator panel'
    sheet['A13'].value = 'DP 2 Breakers' # Subtitles
    sheet['A14'].value = 'DP2-B17 Spare Breaker is Open'
    sheet['A15'].value = 'DP2-B18 ATS-HPB1-C Breaker is Closed'
    sheet['A16'].value = 'DP2-B19 ATS-HPA1-C Breaker is Closed'
    sheet['A17'].value = 'DP2-B20 Spare Breaker is Open'
    sheet['A18'].value = 'DP2-B08 SPD All 3 Green lights (Protected)'
    sheet['A19'].value = 'DP2-B11 ATS-HPB-H Breaker is Closed'
    sheet['A20'].value = 'DP2-B12 CHILLER 2 Breaker is Closed'
    sheet['A21'].value = 'DP2-B13 ATS-HPB-C Breaker is Closed'
    sheet['A22'].value = 'DP2-B14 SPARE Breaker is Open'
    sheet['A23'].value = 'DP2-B15 T-LBP-0 Breaker is Closed'
    sheet['A24'].value = 'DP 1 Breakers' # Subtitles
    sheet['A25'].value = 'DP1-B21 Alt. Feed from MSB3 Breaker is Open'
    sheet['A26'].value = 'DP1-B16 T-PPLL01 Breaker is Closed'
    sheet['A27'].value = 'DP1-B17 ATS-HPD-H Breaker is Closed'
    sheet['A28'].value = 'DP1-B18 ATS-HPC1-C Breaker is Closed'
    sheet['A29'].value = 'DP1-B19 ATS-HPA1-C Breaker is Closed'
    sheet['A30'].value = 'DP1-B20 SPARE Breaker is Open'
    sheet['A31'].value = 'DP1-B08 SPD All 3 Green lights (Protected)'
    sheet['A32'].value = 'DP1-B11 ATS-HPB-H Breaker is Closed'
    sheet['A33'].value = 'DP1-B12 CHILLER 1 Breaker is Closed'
    sheet['A34'].value = 'DP1-B13 ATS-HPA-C Breaker is Closed'
    sheet['A35'].value = 'DP1-B14 ATS-HPA-H Breaker is Closed'
    sheet['A36'].value = 'DP1-B15 T-HLA-0 Breaker is Closed'
    sheet['A37'].value = 'MSB 3' # Subtitles
    sheet['A38'].value = 'MSB3-B12 CI3 Breaker is Closed'
    sheet['A39'].value = 'MSB3-B13 DP3 Breaker is Closed'
    sheet['A40'].value = 'Eaton Xpert meter Events light OFF  (User is X and PW is X)'
    sheet['A41'].value = 'MSB3-B01 Main Breaker is Closed'
    sheet['A42'].value = 'MSB3-B08 SPD All 3 Green lights (Protected)'
    sheet['A43'].value = 'RPTCS (S1 lights are on)'
    sheet['A44'].value = 'Mode Switch is in the Closed Transition position'
    sheet['A45'].value = 'Notes'
    sheet['A45'].alignment = leftTop
    sheet['A45'].font = Font(b=True)

    '''    sheet['A44'].value = 'DP 3 Breakers' # Subtitles
    sheet['A45'].value = 'DP3-B08 SPD All 3 Green lights (Protected)'
    sheet['A46'].value = 'DP3-B12 CHILLER 3 Breaker is'
    sheet['A46'].value = 'DP3-B13 ATS-HPC-C Breaker is'
    sheet['A47'].value = 'DP3-B14 ATS-HPC-H Breaker is'
    sheet['A48'].value = 'DP3-B15 MUA 4 Breaker is'
    sheet['A49'].value = 'DP3-B17 ATS-HPD-H Breaker is' '''

def pg07_engineer_values():
    columns = range(2, 6, 1)
    rowsYes = [18, 31, 40, 42, 43, 44]
    rowsCheck = [3, 8, 10, 11, 12, 14, 15, 16, 17, 19, 20, 21, 22, 23, 24, 26, 27, 28, 29, 30, 32, 33, 34, 35, 36, 38, 39, 41]
    rowsDp = [5]
    rowsT = [9]
    rowsPsi = [4]
    rowsInout = [2]
    rowsAlarm = [6]
    rowsFault = [7]
    rowsBold = []
    # Yes or No values
    for col in columns:
        for row in rowsYes:
            sheet.cell(row=row, column=col).value = 'Yes  /  No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 9, color='696969')
    # ✓ / X values
    for col in columns:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).value = '✓  /  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Temperature
    for col in columns:
        for row in rowsT:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '°F'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # DP
    for col in columns:
        for row in rowsDp:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'DP'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # PSI
    for col in columns:
        for row in rowsPsi:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'PSI'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # In / Out
    for col in columns:
        for row in rowsInout:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = 'In  /  Out'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Alarms
    for col in columns:
        for row in rowsAlarm:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = 'Alarms? Y / N'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Faults
    for col in columns:
        for row in rowsFault:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = 'Faults? Y / N'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Bold text
    for col in columns:
        for row in rowsBold:
            sheet.cell(row=row, column=col).font = Font(bold=True, size=7)
    wb.save('Plymouth_Daily_Rounds.xlsx')


def pg07_colored_cells():
    columnsColor = range(1, 6, 1)
    
    rowsDkGrey = [1]
    for col in columnsColor:
        for row in rowsDkGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx') 
    
    rowsLtGrey = [13, 24, 37]
    for col in columnsColor:
        for row in rowsLtGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Borders
    rows = range(1, 46)
    columns = range(1, 6)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')
