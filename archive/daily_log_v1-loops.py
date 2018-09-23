'''
Loop over this for a row like this: 

for c in range(1, 27): 
        sheet.merge_cells(start_col=c+1, start_row=2, end_col=c+2, end_row=2)
        '''


nl = '\n' 
print(nl)
# sheet.unmerge_cells('A1:A5') # HELP: error unmerge

# imports
import os
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill, NamedStyle


# Set directory
print(os.getcwd())
os.chdir('C:\\Users\\pcurtis7\\Desktop\\_myScripts\\python_excel')
print(os.getcwd(), nl)

# create workbook
wb = openpyxl.Workbook()
print('type(workbook):', type(wb))
print('all.sheetnames:', wb.sheetnames)

# set sheets
sheet = wb.active
sheet.title = 'Daily_Rounds'
s0 = sheet
s1 = wb.create_sheet('Colors')
print('all.sheetnames:', wb.sheetnames)
print('wb.index(s0/Daily_Rounds) =', wb.index(s0))
print('wb.index(s1/Colors) =', wb.index(s1))

# help: how do you change the active sheet?
wb.active = wb.index(s0)
print('sheet/wb.active =', sheet)
wb.active = wb.index(s1)
print('sheet/wb.active =', sheet)
print('s1[1] =', s1[1])

wb.save('daily_log_v2_loops.xlsx')

# General styles
sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=12)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
# wb.add_named_style(highlight) # confirm if this adds to entire workbook
sheet['A1'].style = highlight
sheet['A41'].style = 'highlight'

# Merge Cells
# sheet.merge_cells(start_row=, start_column=, end_row=, end_column=)
# Rows 6-21



for c in range(1, 27): 
        sheet.merge_cells(start_col=c+1, start_row=2, end_col=c+2, end_row=2)







''
start_col = [2, 4, 6, 8]
end_col = [x + 1 for x in start_col]

for i in range (1, 6):
  if i in [1, 4]:
    sheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)
  else:
    for col1, col2 in zip(start_col, end_col):
      sheet.merge_cells(start_row=i, start_column=col1, end_row=i, end_column=col2)
'''
Create a 'start' list with 2, 4, 6, & 8 integers
create an 'end' list that adds +1 for every integer in the 'start' list
For every value (named i) within the range 1 to 6, but not including 6:
  if i is either 1 or 4:
    merge cells starting with row range from 1 to 5, going from columns 1 to 9
  or else:
    for variables col1 & col2 in the zip() function calling start_col & end_col:
      merge cells within row i (range from 1 to 5), going from col1 to col2

I'm not exactly clear on everything here, specifically the objective of the zip() function and what value the col1 and col2 have.

I see the zip() defination but it's still not clear (must be 02:30am!):
  "The zip() function take iterables (zero or more), makes iterator that aggregates elements based on the iterables passed, and returns an iterator of tuples.
  Return Value from zip()
The zip() function returns an iterator of tuples based on the iterable object.
If no parameters are passed, zip() returns an empty iterator
If a single iterable is passed, zip() returns an iterator of 1-tuples. Meaning, the number of elements in each tuple is 1.
If multiple iterables are passed, ith tuple contains ith Suppose, two iterables are passed; one iterable containing 3 and other containing 5 elements. Then, the returned iterator has 3 tuples. It's because iterator stops when shortest iterable is exhaused."

Is my description correct?  What am I missing?
'''




'''
sheet.merge_cells(start_row=6, start_column=2, end_row=6, end_column=3) # SR1 CW Loop
sheet.merge_cells(start_row=6, start_column=4, end_row=6, end_column=5) # SR1 CW Loop
sheet.merge_cells(start_row=6, start_column=6, end_row=6, end_column=7) # SR1 CW Loop
sheet.merge_cells(start_row=6, start_column=8, end_row=6, end_column=9) # SR1 CW Loop
sheet.merge_cells(start_row=10, start_column=2, end_row=10, end_column=3) # PDU 11
sheet.merge_cells(start_row=10, start_column=4, end_row=10, end_column=5) # PDU 11
sheet.merge_cells(start_row=10, start_column=6, end_row=10, end_column=7) # PDU 11
sheet.merge_cells(start_row=10, start_column=8, end_row=10, end_column=9) # PDU 11
sheet.merge_cells(start_row=11, start_column=2, end_row=11, end_column=3) # PDU 9
sheet.merge_cells(start_row=11, start_column=4, end_row=11, end_column=5) # PDU 9
sheet.merge_cells(start_row=11, start_column=6, end_row=11, end_column=7) # PDU 9
sheet.merge_cells(start_row=11, start_column=8, end_row=11, end_column=9) # PDU 9
sheet.merge_cells(start_row=12, start_column=2, end_row=12, end_column=3) # PDU 2
sheet.merge_cells(start_row=12, start_column=4, end_row=12, end_column=5) # PDU 2
sheet.merge_cells(start_row=12, start_column=6, end_row=12, end_column=7) # PDU 2
sheet.merge_cells(start_row=12, start_column=8, end_row=12, end_column=9) # PDU 2
sheet.merge_cells(start_row=13, start_column=2, end_row=13, end_column=3) # PDU 4
sheet.merge_cells(start_row=13, start_column=4, end_row=13, end_column=5) # PDU 4
sheet.merge_cells(start_row=13, start_column=6, end_row=13, end_column=7) # PDU 4
sheet.merge_cells(start_row=13, start_column=8, end_row=13, end_column=9) # PDU 4
sheet.merge_cells(start_row=17, start_column=2, end_row=17, end_column=3) # PDU 1
sheet.merge_cells(start_row=17, start_column=4, end_row=17, end_column=5) # PDU 1
sheet.merge_cells(start_row=17, start_column=6, end_row=17, end_column=7) # PDU 1
sheet.merge_cells(start_row=17, start_column=8, end_row=17, end_column=9) # PDU 1
sheet.merge_cells(start_row=18, start_column=2, end_row=18, end_column=3) # PDU 8
sheet.merge_cells(start_row=18, start_column=4, end_row=18, end_column=5) # PDU 8
sheet.merge_cells(start_row=18, start_column=6, end_row=18, end_column=7) # PDU 8
sheet.merge_cells(start_row=18, start_column=8, end_row=18, end_column=9) # PDU 8
'''
wb.save('daily_log_v2_loops.xlsx')