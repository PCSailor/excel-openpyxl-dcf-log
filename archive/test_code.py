'''
* Test code page for Plymouth_Daily_Rounds.xlsx
    * Template code taken from Page_11

* Page 02 of Plymouth_Daily_Rounds.xlsx
* Server Room 2
* MDF Room
* Fire Pump Room
* VERIFY ALL DATA ENTERED
'''
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook('Plymouth_Daily_Rounds.xlsx')
sheet = wb["test_code"]

print('\nStart next file, \'page_02.py\'')
print('Worksheet list:',  wb.sheetnames) # 
indexNumber = wb.worksheets.index(wb['Page_01'])
print('\'Page_01\' index number = ', indexNumber) # 0
indexNumber = wb.worksheets.index(wb['Page_02'])
print('\'Page_02\' index number = ', indexNumber) # 1
indexNumber = wb.worksheets.index(wb['Page_03'])
print('\'Page_03\' index number = ', indexNumber) # 2
indexNumber = wb.worksheets.index(wb['Page_04'])
print('\'Page_04\' index number = ', indexNumber) # 3
indexNumber = wb.worksheets.index(wb['Page_05'])
print('\'Page_05\' index number = ', indexNumber) # 4
indexNumber = wb.worksheets.index(wb['Page_06'])
print('\'Page_06\' index number = ', indexNumber) # 5
indexNumber = wb.worksheets.index(wb['Page_07'])
print('\'Page_07\' index number = ', indexNumber) # 6
indexNumber = wb.worksheets.index(wb['Page_08'])
print('\'Page_08\' index number = ', indexNumber) # 7
indexNumber = wb.worksheets.index(wb['Page_09'])
print('\'Page_09\' index number = ', indexNumber) # 8
indexNumber = wb.worksheets.index(wb['Page_10'])
print('\'Page_10\' index number = ', indexNumber) # 9
indexNumber = wb.worksheets.index(wb['Page_11'])
print('\'Page_11\' index number = ', indexNumber) # 10
indexNumber = wb.worksheets.index(wb['test_code'])
print('\'test_code\' index number = ', indexNumber) # 11
wb.save('Plymouth_Daily_Rounds.xlsx')

def pg_tc_start():
    print('startup complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg_tc_headers():
    # Local Variables
    # Print Options
    sheet.print_area = 'A1:I51' # TODO: Set print area
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.75
    sheet.page_margins.bottom = 0.75
    sheet.page_margins.header = 0.3
    sheet.page_margins.footer = 0.3
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 12
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 12
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000" # 
    print('headers complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg_tc_merge():
    # Local Variables
    rows = range(1, 51) # FixMe:
    rowsWide = [1, 21, 38, 39, 40, 41, 42] # FixMe:
    # Merges 9 cells into 1 in 1 row
    for row in (1, 30, 38, 39, 47, 48, 49):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    # merge 2 cells into 1 in 1 row
    columns = [(col, col+1) for col in range(2, 9, 2)]
    for row in [6, 7, 8, 9, 13, 18, 19, 23, 25, 26, 27, 28, 29, 31, 32, 33, 35, 36, 37, 40, 44]:
        for col1, col2 in columns:
            sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    # merge 4 cells into 1 cell, all in 1 row
    for row in (10, 21, 35):
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)

    # Column and Row Dimensions
    # Caution: No error is generated if the dimension value does not work
    sheet.column_dimensions['A'].width = 35.00
    for col in ['B', 'D', 'F', 'H']:
        sheet.column_dimensions[col].width = 6.00
    for col in ['C', 'E', 'G', 'I']:
        sheet.column_dimensions[col].width = 12.00
    sheet.row_dimensions[1].height = 20.0
    for row in rows:
        sheet.row_dimensions[row].height = 16.00
    sheet.row_dimensions[1].height = 20.0
    for row in  rowsWide:
        sheet.row_dimensions[row].height = 21.00
    print('merge complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg_tc_namedstyle():
    ''' NamedStyles set (mutable & used when need to apply formatting to different cells at once) '''
    # Local Variables
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='bottom')
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), 
                        right=Side(style='thick'), 
                        top=Side(style='thick'), 
                        bottom=Side(style='thick'))
    # Room Divisions
    sheet['A1'].style = 'rooms'
    sheet['A30'].style = 'rooms'
    sheet['A38'].style = 'rooms'
    #
    # Set Borders
    ''' The merged cell behaves similar to other cell ojects. 
    Its value and format is defined in its top-left cell. 
    In order to change the border of the whole merged cell, change the border of its top-left cell. '''
    rows = range(1, 51)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    print('styles complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg_tc_cell_values():
    # Local Variables
    # Cell values
    sheet['A1'].value = 'Server Room 1'
    sheet['A2'].value = 'Tear off sticky mat (Battery Room)'
    sheet['A3'].value = 'CRAC 24'
    # Server Room #1_Equipment
    sheet['A4'].value = 'Server Room 1'
    sheet['A5'].value = 'Tear off sticky mat (Battery Room)'
    sheet['A6'].value = 'CRAC 24'
    sheet['A4'].value = 'CRAC 23'
    sheet['A5'].value = 'SR1 CHW Loop'
    sheet['A6'].value = 'CRAC 04'
    sheet['A7'].value = 'PDU 11'
    sheet['A8'].value = 'PDU 09'
    sheet['A9'].value = 'PDU 02'
    sheet['A10'].value = 'PDU 04'
    sheet['A11'].value = 'CRAC 26'
    sheet['A12'].value = 'CRAC 05'
    sheet['A13'].value = 'CRAC 06'
    sheet['A14'].value = 'PDU 01'
    sheet['A15'].value = 'PDU 08'
    sheet['A16'].value = 'CRAC 33'
    sheet['A17'].value = 'CRAC 07'
    sheet['A18'].value = 'Humidifier'
    sheet['A19'].value = 'FM 200 (2 tanks)'
    sheet['A20'].value = 'Tear off sticky mat (Hallway)'
    # Server Room #3_Equipment
    sheet['A21'].value = 'Server Room 3'
    sheet['A22'].value = 'Tear of sticky mat (Hallway)'
    sheet['A23'].value = 'CRAC 10'
    sheet['A24'].value = 'CRAC 22'
    sheet['A25'].value = 'CRAC 31'
    sheet['A26'].value = 'PDU 23'
    sheet['A27'].value = 'PDU 22'
    sheet['A28'].value = 'PDU 03'
    sheet['A29'].value = 'PDU 10'
    sheet['A30'].value = 'CRAC 11'
    sheet['A31'].value = 'CRAC 12'
    sheet['A32'].value = 'CRAC 13'
    sheet['A33'].value = 'CRAC 14'
    sheet['A34'].value = 'CRAC 30'
    sheet['A35'].value = 'Humidifier'
    sheet['A36'].value = 'FM 200'
    sheet['A37'].value = 'Tear off sticky mat (Loading Dock)'
    sheet['A38'].value = 'Final Notes:' # StretchGoal: Increase row height, delete comment rows below
    sheet['A39'].value = ''
    sheet['A40'].value = ''
    sheet['A41'].value = ''
    print('cell_values complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg_tc_engineer_values():
    # Local Variables
    columns = [2, 4, 6, 8] # Yes or No values AND ✓ X values
    rows = [2, 20, 22, 37] # Yes or No values
    columnOdd = [3, 5, 7, 9] # RH% AND Hz
    columnEven = [2, 4, 6, 8] # D/P
    rowsCheck = [3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36] # ✓ X values
    rowsHZ = [3, 4, 6, 11, 12, 13, 16, 17, 23, 24, 25, 30, 31, 32, 33, 34] # 
    rowsRH = [18, 35] # 
    rowsDP = [5] # 
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='bottom')
    # Yes or No values
    for col in columns:
        for row in rows:
            sheet.cell(row=row, column=col).value = 'Yes  /  No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 9, color='000000')
    # ✓ X values
    for col in columns:
        for row in rowsCheck:
            # print(col, row)
            sheet.cell(row=row, column=col).value = '✓  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # RH%
    for col in columnOdd:
        for row in rowsRH:
            # print(col, row)
            sheet.cell(row=row, column=col).value = '%RH'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='000000')
    # Hz
    for col in columnOdd:
        for row in rowsHZ:
            # print(col, row)
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='000000')
    # D/P
    for col in columnEven:
        for row in rowsDP:
            # print(col, row)
            sheet.cell(row=row, column=col).value = 'D/P'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='000000')
    print('engineer_values complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg_tc_colored_cells():
    # Local Variables
    rowsColor = [1, 21, 38]
    rowsBlack = [19, 36]
    columnsColor = [1, 2, 4, 6, 8]
    columnsBlack = [2, 6, 8]
    cellBlack = ['B19', 'F19', 'H19', 'B36', 'F36', 'H36']
    for col in columnsColor:
        for row in rowsColor:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    for col in columnsBlack:
        for row in rowsBlack:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
            # stretchGoal: add sheet[cellBlack].value = 'N/A'
    print('colored_cells complete on ', sheet)
    wb.save('Plymouth_Daily_Rounds.xlsx')
























































'''
    # SERVER ROOM 2 CODE BELOW 
# Cell values
sheet['A1'].value = 'Server Room 2'
sheet['A2'].value = 'CRAC 29'
sheet['A3'].value = 'CRAC '
sheet['A4'].value = 'CRAC '
sheet['A5'].value = 'Humidifier'
sheet['A6'].value = 'PDU 21'
sheet['A7'].value = 'PDU 20'
sheet['A8'].value = 'PDU 06'
sheet['A9'].value = 'PDU 14'
sheet['A10'].value = 'FM 200'
sheet['A11'].value = 'CRAC 27'
sheet['A12'].value = 'CRAC 28'
sheet['A13'].value = 'SR2 CHW Loop'
sheet['A14'].value = 'CRAC 17'
sheet['A15'].value = 'CRAC 16'
sheet['A16'].value = 'CRAC 19'
sheet['A17'].value = 'CRAC 18'
sheet['A18'].value = 'PDU 17'
sheet['A19'].value = 'PDU 16'
sheet['A20'].value = 'CRAC 34'
sheet['A21'].value = 'FM 200'
sheet['A22'].value = 'CRAC 15'
sheet['A23'].value = 'CRAC 25'
sheet['A24'].value = 'CRAC 20'
sheet['A25'].value = 'PDU 19'
sheet['A26'].value = 'PDU 18'
sheet['A27'].value = 'PDU 07'
sheet['A28'].value = 'PDU 15'
sheet['A29'].value = 'Tear off Sticky Mat'
sheet['A30'].value = 'MDF'
sheet['A31'].value = 'Tear off Sticky Mat'
sheet['A32'].value = 'PDU 12'
sheet['A33'].value = 'CRAC 08'
sheet['A34'].value = 'Humidifier'
sheet['A35'].value = 'FM 200'
sheet['A36'].value = 'PDU 05'
sheet['A37'].value = 'CRAC 09'
sheet['A38'].value = 'East Battery Room'
sheet['A39'].value = 'Rail 2 Batteries'
sheet['A40'].value = 'CU2 Battery Circuit Breaker'
sheet['A41'].value = 'Eagle Eye Computer Alarms'
sheet['A42'].value = ''
sheet['A43'].value = ''
sheet['A44'].value = 'DC Ground Fault Module reading below 6MA\nPre-alarm=10MA, Alarm=20MA'
sheet['A45'].value = 'Spare Battery Charger'
sheet['A46'].value = ''
sheet['A47'].value = 'Notes:' # StretchGoal: Increase row height, delete comment rows below
sheet['A48'].value = '' # 
sheet['A49'].value = '' # 
sheet['C40'].value = 'Open  /  Closed'
sheet['C41'].value = 'Voltage'
sheet['C42'].value = 'Resistance'
sheet['C43'].value = 'Temerature'
sheet['C44'].value = '✓  X'
sheet['C45'].value = 'Volts'
sheet['C46'].value = 'Amps'

# Engineer Round Values
# Yes or No values
columns = [2, 4, 6, 8]
rows = [29, 31]
# cells = []
for col in columns:
    for row in rows:
        sheet.cell(row=row, column=col).value = 'Yes  /  No'
        sheet.cell(row=row, column=col).alignment = center
        sheet.cell(row=row, column=col).font = Font(size = 8, i=True, color='000000')

# ✓ X values
rowsCheck = [2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 32, 33, 34, 35, 36, 37, 44]
for col in columns:
    for row in rowsCheck:
        # print(col, row)
        sheet.cell(row=row, column=col).value = '✓  X'
        sheet.cell(row=row, column=col).alignment = center
        sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')

# RH%
columnodd = [3, 5, 7, 9]
rowsRH = [5, 34]
for col in columnodd:
    for row in rowsRH:
        # print(col, row)
        sheet.cell(row=row, column=col).value = '%RH'
        sheet.cell(row=row, column=col).alignment = right
        sheet.cell(row=row, column=col).font = Font(size=8, color='000000')

# Hz
rowsHZ = [2, 3, 4, 11, 12, 14, 15, 16, 17, 20, 22, 23, 24, 33, 37]
for col in columnodd:
    for row in rowsHZ:
        # print(col, row)
        sheet.cell(row=row, column=col).value = 'Hz'
        sheet.cell(row=row, column=col).alignment = right
        sheet.cell(row=row, column=col).font = Font(size=8, color='000000')

# Colored Cells
rowscolor = [1, 30, 38, 39, 47]
columnscolor = [1, 2, 4, 6, 8]
for col in columnscolor:
    for row in rowscolor:
        # print(col, row)
        sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')

print('sheet names at end of \'page_02\':', wb.sheetnames)
print('\'page_02\' run with sheet dimensions of ', sheet.dimensions)
wb.save('Plymouth_Daily_Rounds.xlsx')
'''