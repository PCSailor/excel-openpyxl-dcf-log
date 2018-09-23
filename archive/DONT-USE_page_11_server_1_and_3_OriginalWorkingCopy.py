#!/usr/bin/env python3
'''
(openpyxl.worksheet package_Submodules)[https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.html])
(Openpyxl: Working with Styles)[https://openpyxl.readthedocs.io/en/stable/styles.html?highlight=cell%20border#introduction]
(RGB Colors)[https://www.rapidtables.com/web/color/RGB_Color.html]
wb.save('Plymouth_Daily_Rounds.xlsx')
'''
print('\n\'page_11\' is run')
# imports
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook # Todo: Delete once all sheets built
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill, NamedStyle, Color, colors

# Global Variables
nl = '\n'
center = Alignment(horizontal='center', vertical='center')
right = Alignment(horizontal='right', vertical='bottom')
print(nl)
# Set directory
print(os.getcwd())
os.chdir('C:\\Users\\pcurtis7\\Desktop\\_myScripts\\python_excel')
print(os.getcwd(), nl)

# create workbook
wb = openpyxl.Workbook()
print('type(workbook):', type(wb))

# set sheets
sheet = wb.active
sheet.title = 'Plymouth Daily Rounds'
print('all.sheetnames:', wb.sheetnames)
print(sheet, ' is active')

# Print Options
sheet.print_area = 'A1:I43'
sheet.print_options.horizontalCentered = True
sheet.print_options.verticalCentered = True

# Page margins
sheet.page_margins.left = 0.25
sheet.page_margins.right = 0.25
sheet.page_margins.top = 0.75
sheet.page_margins.bottom = 0.75
sheet.page_margins.header = 0.3
sheet.page_margins.footer = 0.3

# Headers & Footers
sheet.oddHeader.center.text = "&[Tab]"
sheet.oddHeader.center.size = 24
sheet.oddHeader.center.font = "Tahoma, Bold"
sheet.oddHeader.center.color = "000000" # 

sheet.oddFooter.left.text = "Page &[Page] of &N"
sheet.oddFooter.left.size = 12
sheet.oddFooter.left.font = "Tahoma, Bold"
sheet.oddFooter.left.color = "000000" # 

sheet.oddFooter.right.text = "&[Path]&[File]"
sheet.oddFooter.right.size = 12
sheet.oddFooter.right.font = "Tahoma, Bold"
sheet.oddFooter.right.color = "000000" # 
wb.save('server_1_and_3.xlsx')

# Merges 9 cells into 1 in 1 row
for row in (1, 4, 24, 41, 42, 43):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
# merge 2 cells into 1 in 1 row
columns = [(col, col+1) for col in range(2, 9, 2)]
for row in [2, 3, 5, 8, 10, 11, 12, 13, 17, 18, 23, 25, 29, 30, 31, 32, 40]:
    for col1, col2 in columns:
        sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
# merge 2 cells into 1 and 4 cells into 1 cell, all in 1 row
for row in (22, 39):
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)


# Column width and Row height
sheet.column_dimensions['A'].width = 30.00
for col in ['B', 'D', 'F', 'H']:
    sheet.column_dimensions[col].width = 4.00
for col in ['C', 'E', 'G', 'I']:
    sheet.column_dimensions[col].width = 10.00
rows = range(1, 43)
for row in rows:
    sheet.row_dimensions[row].Height = 15.00

# Set Named Styles (mutable & used when need to apply formatting to different cells at once)
headerrows = NamedStyle(name='headerrows')
headerrows.font = Font(bold=True, underline='none', sz=12)
headerrows.alignment = center

rooms = NamedStyle(name="rooms")
rooms.font = Font(bold=True, size=12)

rightAlign = NamedStyle(name='rightAlign')
rightAlign.font = Font(b=True, i=True, sz=10)
rightAlign.alignment = Alignment(horizontal='right', vertical='center')

# A1
a1 = sheet['A1'] # Note: When doing rounds be aware for unusual smells, sounds, sights, or anything not normal
a1.style = rooms
a1.font = Font(size=11, i=True, color='FF0000')
a1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
# Header Rows
sheet['A2'].style = headerrows
sheet['A3'].style = 'headerrows'
sheet['B3'].style = 'headerrows'
sheet['C3'].style = 'headerrows'
sheet['D3'].style = 'headerrows'
sheet['F3'].style = 'headerrows'
sheet['H3'].style = 'headerrows'
sheet['A2'].style = rightAlign
sheet['A3'].style = 'rightAlign'
# Room Divisions
sheet['A4'].style = 'rooms'
sheet['E4'].style = 'rooms'
sheet['A24'].style = 'rooms'
sheet['A41'].style = 'rooms'

# BUG: need to fix merged borders not set on merged cells
# Set Borders
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
rows = range(1, 44)
columns = range(1, 10)
for row in rows:
    for col in columns:
        sheet.cell(row, col).border = thin_border
wb.save('server_1_and_3.xlsx')

# Cell values
a1.value = 'Note: When doing rounds be aware for unusual smells, sounds, sights, or anything not normal.'
sheet['A2'].value = 'Engineer Initials:  '
sheet['A3'].value = 'Time of Round:  '
sheet['B3'].value = '02:00'
sheet['D3'].value = '08:00'
sheet['F3'].value = '14:00'
sheet['H3'].value = '20:00'
# Server Room #1_Equipment
sheet['A4'].value = 'Server Room 1'
sheet['A5'].value = 'Tear off Mat (Battery Room)'
sheet['A6'].value = 'CRAC 24'
sheet['A7'].value = 'CRAC 23'
sheet['A8'].value = 'SR1 CHW Loop'
sheet['A9'].value = 'CRAC 04'
sheet['A10'].value = 'PDU 11'
sheet['A11'].value = 'PDU 09'
sheet['A12'].value = 'PDU 02'
sheet['A13'].value = 'PDU 04'
sheet['A14'].value = 'CRAC 26'
sheet['A15'].value = 'CRAC 05'
sheet['A16'].value = 'CRAC 06'
sheet['A17'].value = 'PDU 01'
sheet['A18'].value = 'PDU 08'
sheet['A19'].value = 'CRAC 33'
sheet['A20'].value = 'CRAC 07'
sheet['A21'].value = 'Humidifier'
sheet['A22'].value = 'FM 200 (2 tanks)'
sheet['A23'].value = 'Tear off Mat (Hallway)'

# Server Room #3_Equipment
sheet['A24'].value = 'Server Room 3'
sheet['A25'].value = 'Tear of Mat (Hallway)'
sheet['A26'].value = 'CRAC 10'
sheet['A27'].value = 'CRAC 22'
sheet['A28'].value = 'CRAC 31'
sheet['A29'].value = 'PDU 23'
sheet['A30'].value = 'PDU 22'
sheet['A31'].value = 'PDU 03'
sheet['A32'].value = 'PDU 10'
sheet['A33'].value = 'CRAC 11'
sheet['A34'].value = 'CRAC 12'
sheet['A35'].value = 'CRAC 13'
sheet['A36'].value = 'CRAC 14'
sheet['A37'].value = 'CRAC 30'
sheet['A38'].value = 'Humidifier'
sheet['A39'].value = 'FM 200'
sheet['A40'].value = 'Tear off Mat (Loading Dock Area)'

# Engineer Round Values
# Yes or No values
columns = [2, 4, 6, 8]
rows = [5, 23, 25, 40]
# cells = []
for col in columns:
    for row in rows:
        sheet.cell(row=row, column=col).value = 'Yes  /  No'
        sheet.cell(row=row, column=col).alignment = center
        sheet.cell(row=row, column=col).font = Font(size = 8, i=True, color='000000')

# ✓ X values
rowsCheck = [6, 7, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39]
for col in columns:
    for row in rowsCheck:
        # print(col, row)
        sheet.cell(row=row, column=col).value = '✓  X'
        sheet.cell(row=row, column=col).alignment = center
        sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')

# RH%
columnodd = [3, 5, 7, 9]
rowsRH = [21, 38]
for col in columnodd:
    for row in rowsRH:
        # print(col, row)
        sheet.cell(row=row, column=col).value = '%RH'
        sheet.cell(row=row, column=col).alignment = right
        sheet.cell(row=row, column=col).font = Font(size=8, color='000000')

# Hz
rowsHZ = [6, 7, 9, 14, 15, 16, 19, 20, 26, 27, 28, 33, 34, 35, 36, 37]
for col in columnodd:
    for row in rowsHZ:
        # print(col, row)
        sheet.cell(row=row, column=col).value = 'Hz'
        sheet.cell(row=row, column=col).alignment = right
        sheet.cell(row=row, column=col).font = Font(size=8, color='000000')

# D/P
columnodd = [2, 4, 6, 8]
rowsRH = [8]
for col in columnodd:
    for row in rowsRH:
        # print(col, row)
        sheet.cell(row=row, column=col).value = 'D/P'
        sheet.cell(row=row, column=col).alignment = right
        sheet.cell(row=row, column=col).font = Font(size=8, color='000000')

# Notes
sheet['A41'].value = 'Notes:'

# Colored Cells
rowscolor = [1, 2, 3, 4, 24, 41]
columnscolor = [1, 2, 4, 6, 8]
for col in columnscolor:
    for row in rowscolor:
        # print(col, row)
        sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
        
print('\'page_11\' run with sheet dimensions of ', sheet.dimensions)
wb.save('Plymouth_Daily_Rounds.xlsx')