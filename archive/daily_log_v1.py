#!/usr/bin/env python3
'''
wb.save('daily_log_v1.xlsx')
'''
# sheet.unmerge_cells('A1:A5') # HELP: error unmerge

# imports
import os
import openpyxl
from openpyxl.styles import Font#, Style # HELP: Style not importing
from openpyxl.styles import PatternFill

# Set directory
print(os.getcwd())
os.chdir('C:\\Users\\pcurtis7\\Desktop\\_myScripts\\python_excel')
print(os.getcwd())

# create workbook
wb = openpyxl.Workbook()
wb.save('daily_log_v1.xlsx')
print('type(wb):', type(wb))
print('wb.sheetnames:', wb.sheetnames)
print('wb.active:', wb.active)

# set sheets
sheet = wb.active
print('wb.active:', wb.active)
sheet.title = 'Daily_Rounds'
print('wb.active:', wb.active)
print('sheet.title:', sheet.title)
wb.save('daily_log_v1.xlsx')

# Merge Cells
# sheet.merge_cells(':')
sheet.merge_cells('A1:I1')
sheet.merge_cells('B2:C2')
sheet.merge_cells('D2:E2')
sheet.merge_cells('F2:G2')
sheet.merge_cells('H2:I2')
sheet.merge_cells('B3:C3')
sheet.merge_cells('D3:E3')
sheet.merge_cells('F3:G3')
sheet.merge_cells('H3:I3')
sheet.merge_cells('A4:I4')
sheet.merge_cells('B5:C5')
sheet.merge_cells('D5:E5')
sheet.merge_cells('F5:G5')
sheet.merge_cells('H5:I5')
sheet.merge_cells('B8:C8')
sheet.merge_cells('D8:E8')
sheet.merge_cells('F8:G8')
sheet.merge_cells('H8:I8')
sheet.merge_cells('B10:C10')
sheet.merge_cells('D10:E10')
sheet.merge_cells('F10:G10')
sheet.merge_cells('H10:I10')
sheet.merge_cells('B11:C11')
sheet.merge_cells('D11:E11')
sheet.merge_cells('F11:G11')
sheet.merge_cells('H11:I11')
sheet.merge_cells('B12:C12')
sheet.merge_cells('D12:E12')
sheet.merge_cells('F12:G12')
sheet.merge_cells('H12:I12')
sheet.merge_cells('B13:C13')
sheet.merge_cells('D13:E13')
sheet.merge_cells('F13:G13')
sheet.merge_cells('H13:I13')
sheet.merge_cells('B17:C17')
sheet.merge_cells('D17:E17')
sheet.merge_cells('F17:G17')
sheet.merge_cells('H17:I17')
sheet.merge_cells('B18:C18')
sheet.merge_cells('D18:E18')
sheet.merge_cells('F18:G18')
sheet.merge_cells('H18:I18')
sheet.merge_cells('B22:C22')
sheet.merge_cells('D22:E22')
sheet.merge_cells('F22:G22')
sheet.merge_cells('H22:I22')
sheet.merge_cells('B23:C23')
sheet.merge_cells('D23:E23')
sheet.merge_cells('F23:G23')
sheet.merge_cells('H23:I23')
sheet.merge_cells('A24:I24')
sheet.merge_cells('B25:C25')
sheet.merge_cells('D25:E25')
sheet.merge_cells('F25:G25')
sheet.merge_cells('H25:I25')
sheet.merge_cells('B29:C29')
sheet.merge_cells('D29:E29')
sheet.merge_cells('F29:G29')
sheet.merge_cells('H29:I29')
sheet.merge_cells('B30:C30')
sheet.merge_cells('D30:E30')
sheet.merge_cells('F30:G30')
sheet.merge_cells('H30:I30')
sheet.merge_cells('B31:C31')
sheet.merge_cells('D31:E31')
sheet.merge_cells('F31:G31')
sheet.merge_cells('H31:I31')
sheet.merge_cells('B32:C32')
sheet.merge_cells('D32:E32')
sheet.merge_cells('F32:G32')
sheet.merge_cells('H32:I32')
sheet.merge_cells('B39:C39')
sheet.merge_cells('D39:E39')
sheet.merge_cells('F39:G39')
sheet.merge_cells('H39:I39')
sheet.merge_cells('B40:C40')
sheet.merge_cells('D40:E40')
sheet.merge_cells('F40:G40')
sheet.merge_cells('H40:I40')
sheet.merge_cells('A41:I41')
sheet.merge_cells('A42:I42')
sheet.merge_cells('A43:I43')
wb.save('daily_log_v1.xlsx')

# Header Rows
# sheet[''].value = ''
sheet['A1'].value = 'Note: When doing rounds be aware for unusual smells, sounds, sights, or anything not normal.'
sheet['A2'].value = 'Engineer Initials:'
sheet['B2'].value = 'Int.__________'
sheet['D2'].value = 'Int.__________'
sheet['F2'].value = 'Int.__________'
sheet['H2'].value = 'Int.__________'
sheet['A3'].value = 'Time of Round'
sheet['B3'].value = '02:00'
sheet['D3'].value = '08:00'
sheet['F3'].value = '14:00'
sheet['H3'].value = '20:00'

# Server Room #1_Equipment
sheet['A4'].value = 'Server Room 1'
sheet['A5'].value = 'Tear off Mat (Battery Room)'
sheet['A6'].value = 'CRAC 24'
sheet['A7'].value = 'CRAC 23'
sheet['A8'].value = 'SR1 CHW Loop'
sheet['A9'].value = 'CRAC 04'
sheet['A10'].value = 'PDU 11'
sheet['A11'].value = 'PDU 09'
sheet['A12'].value = 'PDU 02'
sheet['A13'].value = 'PDU 04'
sheet['A14'].value = 'CRAC 26'
sheet['A15'].value = 'CRAC 05'
sheet['A16'].value = 'CRAC 06'
sheet['A17'].value = 'PDU 01'
sheet['A18'].value = 'PDU 08'
sheet['A19'].value = 'CRAC 33'
sheet['A20'].value = 'CRAC 07'
sheet['A21'].value = 'Humidifier'
sheet['A22'].value = 'FM 200 (2 tanks)'
sheet['A23'].value = 'Tear off Mat (Hallway)'


wb.save('daily_log_v1.xlsx')

# Server Room #3_Equipment
sheet['A24'].value = 'Server Room 3'
sheet['A25'].value = 'Tear of Mat (Hallway)'
sheet['A26'].value = 'CRAC 10'
sheet['A27'].value = 'CRAC 22'
sheet['A28'].value = 'CRAC 31'
sheet['A29'].value = 'PDU 23'
sheet['A30'].value = 'PDU 22'
sheet['A31'].value = 'PDU 03'
sheet['A32'].value = 'PDU 10'
sheet['A33'].value = 'CRAC 11'
sheet['A34'].value = 'CRAC 12'
sheet['A35'].value = 'CRAC 13'
sheet['A36'].value = 'CRAC 14'
sheet['A37'].value = 'CRAC 30'
sheet['A38'].value = 'Humidifier'
sheet['A39'].value = 'FM 200'
sheet['A40'].value = 'Tear off Mat (Loading Dock Area)'
wb.save('daily_log_v1.xlsx')

# Engineer Round Values
# Yes or No values
sheet['B5'].value = 'Yes or No'
sheet['D5'].value = 'Yes or No'
sheet['F5'].value = 'Yes or No'
sheet['H5'].value = 'Yes or No'
sheet['B23'].value = 'Yes or No'
sheet['D23'].value = 'Yes or No'
sheet['F23'].value = 'Yes or No'
sheet['H23'].value = 'Yes or No'
sheet['B25'].value = 'Yes or No'
sheet['D25'].value = 'Yes or No'
sheet['F25'].value = 'Yes or No'
sheet['H25'].value = 'Yes or No'
sheet['B40'].value = 'Yes or No'
sheet['D40'].value = 'Yes or No'
sheet['F40'].value = 'Yes or No'
sheet['H40'].value = 'Yes or No'

# ✓ X values
sheet['B6'].value = '✓ X'
sheet['D6'].value = '✓ X'
sheet['F6'].value = '✓ X'
sheet['H6'].value = '✓ X'
sheet['B7'].value = '✓ X'
sheet['D7'].value = '✓ X'
sheet['F7'].value = '✓ X'
sheet['H7'].value = '✓ X'
sheet['B9'].value = '✓ X'
sheet['D9'].value = '✓ X'
sheet['F9'].value = '✓ X'
sheet['H9'].value = '✓ X'
sheet['B10'].value = '✓ X'
sheet['D10'].value = '✓ X'
sheet['F10'].value = '✓ X'
sheet['H10'].value = '✓ X'
sheet['B11'].value = '✓ X'
sheet['D11'].value = '✓ X'
sheet['F11'].value = '✓ X'
sheet['H11'].value = '✓ X'
sheet['B12'].value = '✓ X'
sheet['D12'].value = '✓ X'
sheet['F12'].value = '✓ X'
sheet['H12'].value = '✓ X'
sheet['B13'].value = '✓ X'
sheet['D13'].value = '✓ X'
sheet['F13'].value = '✓ X'
sheet['H13'].value = '✓ X'
sheet['B14'].value = '✓ X'
sheet['D14'].value = '✓ X'
sheet['F14'].value = '✓ X'
sheet['H14'].value = '✓ X'
sheet['B15'].value = '✓ X'
sheet['D15'].value = '✓ X'
sheet['F15'].value = '✓ X'
sheet['H15'].value = '✓ X'
sheet['B16'].value = '✓ X'
sheet['D16'].value = '✓ X'
sheet['F16'].value = '✓ X'
sheet['H16'].value = '✓ X'
sheet['B17'].value = '✓ X'
sheet['D17'].value = '✓ X'
sheet['F17'].value = '✓ X'
sheet['H17'].value = '✓ X'
sheet['B18'].value = '✓ X'
sheet['D18'].value = '✓ X'
sheet['F18'].value = '✓ X'
sheet['H18'].value = '✓ X'
sheet['B19'].value = '✓ X'
sheet['D19'].value = '✓ X'
sheet['F19'].value = '✓ X'
sheet['H19'].value = '✓ X'
sheet['B20'].value = '✓ X'
sheet['D20'].value = '✓ X'
sheet['F20'].value = '✓ X'
sheet['H20'].value = '✓ X'
sheet['B21'].value = '✓ X'
sheet['D21'].value = '✓ X'
sheet['F21'].value = '✓ X'
sheet['H21'].value = '✓ X'
sheet['B26'].value = '✓ X'
sheet['D26'].value = '✓ X'
sheet['F26'].value = '✓ X'
sheet['H26'].value = '✓ X'
sheet['B27'].value = '✓ X'
sheet['D27'].value = '✓ X'
sheet['F27'].value = '✓ X'
sheet['H27'].value = '✓ X'
sheet['B28'].value = '✓ X'
sheet['D28'].value = '✓ X'
sheet['F28'].value = '✓ X'
sheet['H28'].value = '✓ X'
sheet['B29'].value = '✓ X'
sheet['D29'].value = '✓ X'
sheet['F29'].value = '✓ X'
sheet['H29'].value = '✓ X'
sheet['B30'].value = '✓ X'
sheet['D30'].value = '✓ X'
sheet['F30'].value = '✓ X'
sheet['H30'].value = '✓ X'
sheet['B31'].value = '✓ X'
sheet['D31'].value = '✓ X'
sheet['F31'].value = '✓ X'
sheet['H31'].value = '✓ X'
sheet['B32'].value = '✓ X'
sheet['D32'].value = '✓ X'
sheet['F32'].value = '✓ X'
sheet['H32'].value = '✓ X'
sheet['B33'].value = '✓ X'
sheet['D33'].value = '✓ X'
sheet['F33'].value = '✓ X'
sheet['H33'].value = '✓ X'
sheet['B34'].value = '✓ X'
sheet['D34'].value = '✓ X'
sheet['F34'].value = '✓ X'
sheet['H34'].value = '✓ X'
sheet['B35'].value = '✓ X'
sheet['D35'].value = '✓ X'
sheet['F35'].value = '✓ X'
sheet['H35'].value = '✓ X'
sheet['B36'].value = '✓ X'
sheet['D36'].value = '✓ X'
sheet['F36'].value = '✓ X'
sheet['H36'].value = '✓ X'
sheet['B37'].value = '✓ X'
sheet['D37'].value = '✓ X'
sheet['F37'].value = '✓ X'
sheet['H37'].value = '✓ X'
sheet['B38'].value = '✓ X'
sheet['D38'].value = '✓ X'
sheet['F38'].value = '✓ X'
sheet['H38'].value = '✓ X'
sheet['D22'].value = '✓ X'
sheet['D39'].value = '✓ X'

# RH%
sheet['C21'].value = 'RH%'
sheet['E21'].value = 'RH%'
sheet['G21'].value = 'RH%'
sheet['I21'].value = 'RH%'
sheet['C38'].value = 'RH%'
sheet['E38'].value = 'RH%'
sheet['G38'].value = 'RH%'
sheet['I38'].value = 'RH%'

# D/P
sheet['B8'].value = 'D/P'
sheet['D8'].value = 'D/P'
sheet['F8'].value = 'D/P'
sheet['H8'].value = 'D/P'

# Hz
sheet['C6'].value = 'Hz'
sheet['E6'].value = 'Hz'
sheet['G6'].value = 'Hz'
sheet['I6'].value = 'Hz'
sheet['C7'].value = 'Hz'
sheet['E7'].value = 'Hz'
sheet['G7'].value = 'Hz'
sheet['I7'].value = 'Hz'
sheet['C9'].value = 'Hz'
sheet['E9'].value = 'Hz'
sheet['G9'].value = 'Hz'
sheet['I9'].value = 'Hz'
sheet['C14'].value = 'Hz'
sheet['E14'].value = 'Hz'
sheet['G14'].value = 'Hz'
sheet['I14'].value = 'Hz'
sheet['C15'].value = 'Hz'
sheet['E15'].value = 'Hz'
sheet['G15'].value = 'Hz'
sheet['I15'].value = 'Hz'
sheet['C16'].value = 'Hz'
sheet['E16'].value = 'Hz'
sheet['G16'].value = 'Hz'
sheet['I16'].value = 'Hz'
sheet['C19'].value = 'Hz'
sheet['E19'].value = 'Hz'
sheet['G19'].value = 'Hz'
sheet['I19'].value = 'Hz'
sheet['C20'].value = 'Hz'
sheet['E20'].value = 'Hz'
sheet['G20'].value = 'Hz'
sheet['I20'].value = 'Hz'
sheet['C26'].value = 'Hz'
sheet['E26'].value = 'Hz'
sheet['G26'].value = 'Hz'
sheet['I26'].value = 'Hz'
sheet['C27'].value = 'Hz'
sheet['E27'].value = 'Hz'
sheet['G27'].value = 'Hz'
sheet['I27'].value = 'Hz'
sheet['C28'].value = 'Hz'
sheet['E28'].value = 'Hz'
sheet['G28'].value = 'Hz'
sheet['I28'].value = 'Hz'
sheet['C33'].value = 'Hz'
sheet['E33'].value = 'Hz'
sheet['G33'].value = 'Hz'
sheet['I33'].value = 'Hz'
sheet['C34'].value = 'Hz'
sheet['E34'].value = 'Hz'
sheet['G34'].value = 'Hz'
sheet['I34'].value = 'Hz'
sheet['C35'].value = 'Hz'
sheet['E35'].value = 'Hz'
sheet['G35'].value = 'Hz'
sheet['I35'].value = 'Hz'
sheet['C36'].value = 'Hz'
sheet['E36'].value = 'Hz'
sheet['G36'].value = 'Hz'
sheet['I36'].value = 'Hz'
sheet['C37'].value = 'Hz'
sheet['E37'].value = 'Hz'
sheet['G37'].value = 'Hz'
sheet['I37'].value = 'Hz'

# Notes
sheet['A41'].value = 'Notes:'

wb.save('daily_log_v1.xlsx')

# Styling
# Samples
# Darkest to lightest
sheet['K3'].value = '696969'
sheet['K2'].value = '808080'
sheet['K4'].value = 'A9A9A9'
sheet['K5'].value = 'C0C0C0'
# these two look beige
sheet['K1'].value = 'D3D3D3'
sheet['K6'].value = 'DCDCDC'

sheet['J1'].fill = PatternFill(fgColor='D3D3D3', fill_type = 'solid')
sheet['J2'].fill = PatternFill(fgColor='808080', fill_type = 'solid')
sheet['J3'].fill = PatternFill(fgColor='696969', fill_type = 'solid')
sheet['J4'].fill = PatternFill(fgColor='A9A9A9', fill_type = 'solid')
sheet['J5'].fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
sheet['J6'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
# End

# Active Cells
sheet['A1'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
#A2
sheet['A2'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
sheet['B2'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
sheet['D2'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
sheet['F2'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
sheet['H2'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
# A3
sheet['A3'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
sheet['B3'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
sheet['D3'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
sheet['F3'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
sheet['H3'].fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')

sheet['A4'].fill = PatternFill(fgColor='808080', fill_type = 'solid')
sheet['A24'].fill = PatternFill(fgColor='808080', fill_type = 'solid')
sheet['A41'].fill = PatternFill(fgColor='A9A9A9', fill_type = 'solid')

wb.save('daily_log_v1.xlsx')