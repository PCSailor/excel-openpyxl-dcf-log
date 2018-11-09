#!/usr/bin/env python3
'''
* 
wb.save('Plymouth_Daily_Rounds.xlsx')
'''
print('\nStart next file, \'page_10\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_10"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
columns = [(col, col+1) for col in range(2, 6, 2)]
center = Alignment(horizontal='center', vertical='center')
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# Local Variables
values = ['UPS Room 1', 'West Battery Room Hydrogen Monitor Levels', 'MSB 1', 'RPTCS (S1 lights are on)', 'Mode Switch is in the Closed Transition position', 'Eaton Xpert meter Events light OFF (User is X and PW is X)', 'MSB1-B01 (Main) Breaker is Closed', 'MSB1-B08 SPD All 3 Green lights (Protected)', 'MSB1-B12 DP1 Breaker is Closed', 'MSB1-B14 Spare (de-powered)', 

'CI 1', 'CI1-B05 (CU1-M1 Maint. Bypass) Breaker is Closed', 'CI1-B11 (CU1-M1 Input) Breaker is Closed', 'CI1-B06 (CU1-M1 Static Bypass) Breaker is Closed', '** Ensure key is in locked position before touching STS display screen!! **', 'STS1A is on preferred source 1', 

'CC 1', 'CC1-B99 (LBB) Breaker is Open', 'CC1-B01 (MIB) Breaker is Closed', 'CC1-B05 (MBB) Breaker is Open', 

'CO 1', 'Eaton Xpert meter Events light OFF (User is X and PW is X)', 'CO1-B01 (Isolation switch) Breaker is Closed', 'STS1B is on preferred source 1', 'CO1-B13 (STS2A) Breaker is Closed', 'CO1-B14 (STS3A) Breaker is Closed', 'CO1-B15 (STS2C) Breaker is Closed', 'CO1-B16 (Spare) Breaker is Open', 'Eaton Breaker Interface Module II Alarm light OFF', 'CO1-B18 (Spare) Breaker is Open', 'CO1-B11 (STS1A) Breaker is Closed', 'CO1-B12 (STS1B) Breaker is Closed', 'CO1-B17 (Spare) Breaker is Open', 'West Electrical Rm. Leak Detection', 'CRAC 38', 'CU1-M1 (UPS 1)', '', '', 'CU1-M1 (UPS 1) MBB Module Battery Breaker is Closed', 'CRAC 39', 

'Rail 1 Batteries', 'Eagle Eye Computer Alarms', '', '', 'DC Ground Fault Module below 6MA\n(Pre-alarm = 10MA, Alarm = 20MA)', 'CU1 Battery Breaker is Closed', 'Room Temperature', 

'Notes']

def pg10_headers():
    # Print Options
    sheet.print_area = 'A1:I48' # Todo: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 7
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg10_merge():
    columns = [(col, col+1) for col in range(2, 10, 2)]
    # merge 8 cells in 1 Row
    sheet.merge_cells(start_row=10, start_column=2, end_row=10, end_column=9)
    # merge 3 cells in 1 Column
    sheet.merge_cells(start_row=36, start_column=1, end_row=38, end_column=1)
    sheet.merge_cells(start_row=42, start_column=1, end_row=44, end_column=1)
    # Merges 9 cells into 1 in 1 row
    for row in [1, 3, 11, 15, 17, 21, 41, 48]:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    # merge 2 cells into 1 in 1 row
    for row in [2, 4, 5, 6, 7, 8, 9, 12, 13, 14, 16, 18, 19, 20, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 36, 37, 38, 39, 41, 42, 43, 44, 45, 46, 47, 48]:
            for col1, col2 in columns:
                sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Dimensions_Columns & Rows
    sheet.column_dimensions['A'].width = 51.25
    for col in ['B', 'D', 'F', 'H']:
        sheet.column_dimensions[col].width = 4.50
    for col in ['C', 'E', 'G', 'I']:
        sheet.column_dimensions[col].width = 8.00
    rows = range(1, 50)
    for row in rows:
        sheet.row_dimensions[row].height = 14.50
    sheet.row_dimensions[48].height = 30.00
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Page Font
    rows = range(1, 49)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).font = Font(size = 10, i=False, color='000000')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg10_namedstyle():
    # Styles
    sheet['A1'].style = 'rooms'
    sheet['A41'].style = 'rooms'
    sheet['A3'].style = 'subtitles'
    sheet['A11'].style = 'subtitles'
    sheet['A17'].style = 'subtitles'
    sheet['A21'].style = 'subtitles'
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg10_cell_values():
    # Cell values
    def populate(sheet, col, row, values):
        row = int(row)
        for i, value in enumerate(values):
            sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list
    #
    # sheet.cell(row=8, column=2).fill = PatternFill(fgColor='000000', fill_type = 'solid')
    sheet['A15'].font = Font(size=12, b=True, i=True, color='FF0000')
    sheet['A15'].alignment = center
    sheet.cell(row=36, column=1).alignment = left
    sheet.cell(row=42, column=1).alignment = left
    sheet['A48'].alignment = leftTop
    sheet['A48'].font = Font(b=True)

def pg10_engineer_values():
    columns = range(2, 10, 1)
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    rowsCheck = [4, 5, 6, 7, 8, 9, 12, 13, 14, 16, 18, 19, 20, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 39, 45, 46]
    rowsT = [44, 47]
    rowsR = [43]
    rowsHz = [35, 40]
    rowsVdc = [2, 42]
    rowEf = [2]
    rowsKw = [36]
    rowsKva = [37]
    rowsAct = [38]
    # ✓ / X values
    for col in columns:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).value = '✓  /  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # ✓  /  X on Even
    for col in columnEven:
        for row in rowsHz:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = '✓ / X'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # Hz on Odd
    for col in columnOdd:
        for row in rowsHz:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # vDC
    for col in columnEven:
        for row in rowsVdc:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'vDC'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Resistance
    for col in columnEven:
        for row in rowsR:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'Ω'
            sheet.cell(row=row, column=col).font = Font(size=9, color='696969')
    # Temperature
    for col in columnEven:
        for row in rowsT:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '°F'
            sheet.cell(row=row, column=col).font = Font(size=9, color='696969')
    # kW-Out
    for col in columnEven:
        for row in rowsKw:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'kW->'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # kVA-Out
    for col in columnEven:
        for row in rowsKva:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'kVA->'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Active Events
    for col in columnEven:
        for row in rowsAct:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = 'Active Events?'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # EF%
    for col in columnEven:
        for row in rowEf:
            sheet.cell(row=row, column=col).alignment = ctrdwn
            sheet.cell(row=row, column=col).value = 'EF4%    /    EF5%'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg10_colored_cells():
    columnsColor = range(1, 10, 1)
    rowsDkGrey = [1, 41]
    for col in columnsColor:
        for row in rowsDkGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    rowsLtGrey = [3, 11, 17, 21]
    for col in columnsColor:
        for row in rowsLtGrey:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    
    sheet.cell(row=10, column=2).fill = PatternFill(fgColor='000000', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')

    # Borders
    rows = range(1, 49)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')
