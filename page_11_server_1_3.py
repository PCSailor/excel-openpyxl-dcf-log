#! python3
'''
* Page 11 of Plymouth_Daily_Rounds.xlsx
* Server Room #1
* Server Room #3
'''
# Imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill, GradientFill
wb = load_workbook('Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_11"]
print('\nStart next file, \'page_11\'')
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')

# Global Variables
center = Alignment(horizontal='center', vertical='center')
ctrdwn = Alignment(horizontal='center', vertical='bottom')
right = Alignment(horizontal='right', vertical='bottom')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
leftTop = Alignment(horizontal='left', vertical='top')
wrap = Alignment(wrap_text=True)
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
rows = range(1, 39)
# Local Variable
''' values = ['Server Room 1', 'Tear off sticky mat (Battery Room)'
    sheet['A3'].value = 'CRAC 24'
    sheet['A4'].value = 'Server Room 1' # Server Room #1_Equipment
    sheet['A5'].value = 'Tear off sticky mat (Battery Room)'
    sheet['A6'].value = 'CRAC 24'
    sheet['A4'].value = 'CRAC 23'
    sheet['A5'].value = 'SR1 CHW Loop'
    sheet['A6'].value = 'CRAC 04'
    sheet['A7'].value = 'PDU 11'
    sheet['A8'].value = 'PDU 09'
    sheet['A9'].value = 'PDU 02'
    sheet['A10'].value = 'PDU 04'
    sheet['A11'].value = 'CRAC 26'
    sheet['A12'].value = 'CRAC 05'
    sheet['A13'].value = 'CRAC 06'
    sheet['A14'].value = 'PDU 01'
    sheet['A15'].value = 'PDU 08'
    sheet['A16'].value = 'CRAC 33'
    sheet['A17'].value = 'CRAC 07'
    sheet['A18'].value = 'Humidifier'
    sheet['A19'].value = 'FM 200 (2 tanks)'
    sheet['A20'].value = 'Tear off sticky mat (Hallway)'
    sheet['A21'].value = 'Server Room 3' # Server Room #3_Equipment
    sheet['A22'].value = 'Tear of sticky mat (Hallway)'
    sheet['A23'].value = 'CRAC 10'
    sheet['A24'].value = 'CRAC 22'
    sheet['A25'].value = 'CRAC 31'
    sheet['A26'].value = 'PDU 23'
    sheet['A27'].value = 'PDU 22'
    sheet['A28'].value = 'PDU 03'
    sheet['A29'].value = 'PDU 10'
    sheet['A30'].value = 'CRAC 11'
    sheet['A31'].value = 'CRAC 12'
    sheet['A32'].value = 'CRAC 13'
    sheet['A33'].value = 'CRAC 14'
    sheet['A34'].value = 'CRAC 30'
    sheet['A35'].value = 'Humidifier'
    sheet['A36'].value = 'FM 200'
    sheet['A37'].value = 'Tear off sticky mat (Loading Dock)'
    sheet['A38'].value = 'Final Notes:'

]
'''

def pg11_headers():
    # Print Options
    sheet.print_area = 'A1:I38' # TODO: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.8
    sheet.page_margins.bottom = 0.75
    sheet.page_margins.header = 0.55
    sheet.page_margins.footer = 0.3
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 12
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 6
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg11_merge():
    columns = [(col, col+1) for col in range(2, 9, 2)] # Todo: Move?
    # Merge 9 cells into 1 in 1 row
    for row in (1, 21, 38, 39, 40, 41, 42, 43):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    # Merge 2 cells into 1 in 1 row
    for row in [2, 5, 7, 8, 9, 10, 14, 15, 20, 22, 26, 27, 28, 29, 37]:
        for col1, col2 in columns:
            sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    # merge 2 cells into 1 and 4 cells into 1 cell, all in 1 row
    for row in (19, 36):
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)

    # Dimensions_Columns & Rows
    # Caution: No error is generated if the dimension value does not work
    # rows = range(1, 37) # Todo: Move?
    sheet.column_dimensions['A'].width = 36.00
    for col in ['B', 'D', 'F', 'H']:
        sheet.column_dimensions[col].width = 6.00
    for col in ['C', 'E', 'G', 'I']:
        sheet.column_dimensions[col].width = 10.00
    for row in rows:
        sheet.row_dimensions[row].height = 15.00
    sheet.row_dimensions[38].height = 70.00
    wb.save('Plymouth_Daily_Rounds.xlsx')
     
    # Page Fonts
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).font = Font(size = 10, i=False, color='000000') # changes entire sheet
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg11_namedstyle():
    ''' NamedStyles set (mutable & used when need to apply formatting to different cells at once) '''
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    # Room Divisions
    sheet['A1'].style = 'rooms'
    sheet['A21'].style = 'rooms'

    sheet['A38'].alignment = leftTop
    sheet['A38'].font = Font(b=True)

def pg11_cell_values():
    # Cell values
    ''' def populate(sheet, col, row, values):
            row = int(row)
            for i, value in enumerate(values):
                    sheet['{}{}'.format(col, row+i)].value = value
    populate(sheet, 'A', 1, values) # worksheet, start column, start row, data list '''
    #
    sheet['A1'].value = 'Server Room 1'
    sheet['A2'].value = 'Tear off sticky mat (Battery Room)'
    sheet['A3'].value = 'CRAC 24'
    sheet['A4'].value = 'Server Room 1' # Server Room #1_Equipment
    sheet['A5'].value = 'Tear off sticky mat (Battery Room)'
    sheet['A6'].value = 'CRAC 24'
    sheet['A4'].value = 'CRAC 23'
    sheet['A5'].value = 'SR1 CHW Loop'
    sheet['A6'].value = 'CRAC 04'
    sheet['A7'].value = 'PDU 11'
    sheet['A8'].value = 'PDU 09'
    sheet['A9'].value = 'PDU 02'
    sheet['A10'].value = 'PDU 04'
    sheet['A11'].value = 'CRAC 26'
    sheet['A12'].value = 'CRAC 05'
    sheet['A13'].value = 'CRAC 06'
    sheet['A14'].value = 'PDU 01'
    sheet['A15'].value = 'PDU 08'
    sheet['A16'].value = 'CRAC 33'
    sheet['A17'].value = 'CRAC 07'
    sheet['A18'].value = 'Humidifier'
    sheet['A19'].value = 'FM 200 (2 tanks)'
    sheet['A20'].value = 'Tear off sticky mat (Hallway)'
    sheet['A21'].value = 'Server Room 3' # Server Room #3_Equipment
    sheet['A22'].value = 'Tear of sticky mat (Hallway)'
    sheet['A23'].value = 'CRAC 10'
    sheet['A24'].value = 'CRAC 22'
    sheet['A25'].value = 'CRAC 31'
    sheet['A26'].value = 'PDU 23'
    sheet['A27'].value = 'PDU 22'
    sheet['A28'].value = 'PDU 03'
    sheet['A29'].value = 'PDU 10'
    sheet['A30'].value = 'CRAC 11'
    sheet['A31'].value = 'CRAC 12'
    sheet['A32'].value = 'CRAC 13'
    sheet['A33'].value = 'CRAC 14'
    sheet['A34'].value = 'CRAC 30'
    sheet['A35'].value = 'Humidifier'
    sheet['A36'].value = 'FM 200'
    sheet['A37'].value = 'Tear off sticky mat (Loading Dock)'
    sheet['A38'].value = 'Final Notes:'
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg11_engineer_values():
    # Local Variables
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]
    rowYes = [2, 20, 22, 37] # Yes or No values
    rowsCheck = [3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35] # ✓ X values
    rowsHZ = [3, 4, 6, 11, 12, 13, 16, 17, 23, 24, 25, 30, 31, 32, 33, 34] # 
    rowsRH = [18, 35] # 
    rowsDP = [5]
    # Yes or No values
    for col in columnEven:
        for row in rowYes:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = 'Yes    /    No'
            sheet.cell(row=row, column=col).font = Font(size = 9, color='696969')
    # ✓ X values
    for col in columnEven:
        for row in rowsCheck:
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).value = '✓   X'
            sheet.cell(row=row, column=col).font = Font(size=8, color='DCDCDC')
    # RH%
    for col in columnOdd:
        for row in rowsRH:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = '%RH'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # Hz
    for col in columnOdd:
        for row in rowsHZ:
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    # D/P
    for col in columnEven:
        for row in rowsDP:
            sheet.cell(row=row, column=col).value = 'D/P'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='696969')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg11_colored_cells():
    # cellBlack = ['B19', 'F19', 'H19', 'B36', 'F36', 'H36']
    rowsColor = [1, 21]
    rowsBlack = [19, 36]
    columnsColor = [1, 2, 4, 6, 8]
    columnsBlack = [2, 6, 8]
    for col in columnsColor:
        for row in rowsColor:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    for col in columnsBlack:
        for row in rowsBlack:
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')
    wb.save('Plymouth_Daily_Rounds.xlsx')
     
    # Borders
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    wb.save('Plymouth_Daily_Rounds.xlsx')