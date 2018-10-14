#! python3
'''
* East UPS Room
* Fire Pump Room
* Loading Dock Area
* Mechanical Room
'''
print('Start next file, \'page_04\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Page_04"]
print('Active sheet is ', sheet)
print('04-01')
wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_headers():
    # center = Alignment(horizontal='center', vertical='center')
    # right = Alignment(horizontal='right', vertical='bottom')
    # Print Options
    sheet.print_area = 'A1:I42' # TODO: set cell region
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    # Page margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.25
    sheet.page_margins.footer = 0.25
    # Headers & Footers
    sheet.oddHeader.center.text = "&[File]"
    sheet.oddHeader.center.size = 20
    sheet.oddHeader.center.font = "Tahoma, Bold"
    sheet.oddHeader.center.color = "000000" # 
    sheet.oddFooter.left.text = "&[Tab] of 11"
    sheet.oddFooter.left.size = 10
    sheet.oddFooter.left.font = "Tahoma, Bold"
    sheet.oddFooter.left.color = "000000" # 
    sheet.oddFooter.right.text = "&[Path]&[File]"
    sheet.oddFooter.right.size = 6
    sheet.oddFooter.right.font = "Tahoma, Bold"
    sheet.oddFooter.right.color = "000000"
    print('04-02')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_merge():
    # center = Alignment(horizontal='center', vertical='center')
    # right = Alignment(horizontal='right', vertical='bottom')
    # Merges 9 cells into 1 in 1 row
    for row in (1, 5, 12, 13, 23, 24):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    # merge 2 cells into 1 in 1 row
    columns = [(col, col+1) for col in range(2, 9, 2)]
    for row in [2, 3, 4, 6, 7, 8, 9, 10, 11, 15, 16, 17, 18, 19, 20, 21, 22, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34]:
        for col1, col2 in columns:
            sheet.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    # Column width and Row height
    sheet.column_dimensions['A'].width = 30.00
    for col in ['B', 'D', 'F', 'H']:
        sheet.column_dimensions[col].width = 4.00
    for col in ['C', 'E', 'G', 'I']:
        sheet.column_dimensions[col].width = 10.00
    rows = range(1, 43)
    for row in rows:
        sheet.row_dimensions[row].Height = 15.00
    # Wrap text Column A
    rows = range(1, 31)
    for row in rows:
        for col in columns:
            sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
    sheet.merge_cells(start_row=30, start_column=1, end_row=32, end_column=1)
    print('04-03')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_namedstyle():
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), 
                        right=Side(style='thick'), 
                        top=Side(style='thick'), 
                        bottom=Side(style='thick'))
    # Styles
    sheet['A1'].style = 'rooms'
    sheet['A12'].style = 'rooms'
    sheet['A24'].style = 'rooms'
    sheet['A28'].style = 'rooms'
    '''    sheet['B21'].style = 'rightAlign' # Todo: Add into forLoop
    sheet['B24'].style = 'rightAlign'
    sheet['B25'].style = 'rightAlign'
    sheet['B27'].style = 'rightAlign' '''
    sheet.cell(row=30, column=1).alignment = center
    sheet['A5'].alignment = center

    # Borders
    rows = range(1, 80)
    columns = range(1, 10)
    for row in rows:
        for col in columns:
            sheet.cell(row, col).border = thin_border
    print('04-04')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_cell_values():
    # Cell values
    sheet['A1'].value = 'CC3'
    sheet['A2'].value = 'CC3-B05 (MBB) Breaker is Open'
    sheet['A3'].value = 'CC3-B01 (MIB) Breaker is Closed'
    sheet['A4'].value = 'CC3-B99 (LBB) Breaker is Open'
    sheet['A5'].value = 'Ensure Key is in Locked position before touching STS screen'
    sheet['A6'].value = 'STS3A is on preferred source 1'
    sheet['A7'].value = 'STS3B is on preferred source 1'
    sheet['A8'].value = 'EF 4'
    sheet['A9'].value = 'EF 5'
    sheet['A10'].value = 'East Electrical Room Leak Detection'
    sheet['A11'].value = 'Tear off sticky mat for SR3 (East side)'
    sheet['A12'].value = 'Fire Pump/ Pre action Room' # Note: Room

    sheet['A13'].value = 'Only on the 20:00 rounds check pre action valves to make sure they’re open (if open put a check next to each zone):'
    sheet['A14'].value = 'Zone 1'
    sheet['B14'].value = 'Zone 2'
    sheet['C14'].value = 'Zone 3'
    sheet['D14'].value = 'Zone 4'
    sheet['E14'].value = 'Zone 5'
    sheet['F14'].value = 'Zone 6'
    sheet['G14'].value = 'Zone 7'
    sheet['H14'].value = 'Wet system level 1-4 '
    sheet['I14'].value = 'Wet system level 0 (Corridors)'

    sheet['A15'].value = 'Jockey pump controller in Auto '
    sheet['A16'].value = 'Fire pump controller in Auto '
    sheet['A17'].value = 'Fire pump is on Normal source power'
    sheet['A18'].value = 'System water pressure left side of controller (140 -150psi)'
    sheet['A19'].value = 'System Nitorgen PSI (inside the red cabinet)'
    sheet['A20'].value = 'At Nitrogen tank regulator: (Replace with Extra Dry Nitrogen at 200PSI)'
    sheet['A21'].value = 'Main building water meter (Total) readings (Top reading)'
    sheet['A22'].value = 'Is Building Main-Drain Water Leaking?'
    sheet['A23'].value = 'If drain pipe has water leaking, check the air-bleed-off-valve in the penthouse stairwell for leaks.'
    sheet['A24'].value = 'Loading Dock Area' # Note: Room
    sheet['A25'].value = 'Do we need to order salt? If yes let the Chief Engineer know.'
    sheet['A26'].value = 'Check brine level (should be at the indicating line).'
    sheet['A27'].value = 'HP LL- 5 Ok  (Fan is ok, If there\'s sweating of pipes check operation of HP)'
    sheet['A28'].value = 'Mechanical / Chill Water Units Room' # Note: Room
    sheet['A29'].value = 'Cooling Twr. Supply  water meter reading.'
    sheet['A30'].value = 'Write down the water softener gallon readings from each softener.'
    # sheet['A31'].value = '' # Todo: merge with line 29
    # sheet['A32'].value = '' # Todo: merge with line 29
    sheet['A33'].value = 'Well meter reading'
    sheet['A34'].value = 'HP LL- 4 Ok  (Fan is ok, If there\'s sweating of pipes check operation of HP)'
    sheet['A35'].value = 'CHWP #3'
    sheet['A36'].value = 'CHWP #5'
    sheet['A37'].value = 'CHWP #2'
    sheet['A38'].value = 'CHWP #4'
    sheet['A39'].value = 'CHWP #1'
    sheet['A40'].value = 'CDW to CHW makeup' # Todo: two line 40's
    # sheet['A40'].value = 'CHW'
    sheet['A41'].value = 'CHW Filter PSI (23psi)'
    sheet['A42'].value = 'Bladder tank pressure (<30)'
    sheet['A43'].value = 'CHW Lakos Bag filter'
    sheet['A44'].value = 'Condenser Supply Temp.  East Side (68 – 85)'
    sheet['A45'].value = 'CWP-6 VFD'
    sheet['A46'].value = 'CWP-1 VFD'
    sheet['A47'].value = 'CWP-4 VFD'
    sheet['A48'].value = 'CWP-3 VFD'
    sheet['A49'].value = 'CDWF  VFD '
    sheet['A50'].value = 'CWP-2 VFD'
    sheet['A51'].value = 'CWP-5 VFD'
    sheet['A52'].value = 'TWR Fan- 6 VFD'
    sheet['A53'].value = 'TWR Fan- 5 VFD'
    sheet['A54'].value = 'CHWR Header Temp East'
    sheet['A55'].value = 'CHWR Temp (Bypass) East'
    sheet['A56'].value = 'Lakos Separator (6psi)'
    sheet['A57'].value = 'CHWS Temp East'
    sheet['A58'].value = 'CHWP #3 VFD'
    sheet['A59'].value = 'Well VFD'
    sheet['A60'].value = 'CHWP #2 VFD'
    sheet['A61'].value = 'CHWP #4 VFD'
    sheet['A62'].value = 'CHWP #1 VFD'
    sheet['A63'].value = 'CHWP #5 VFD'
    sheet['A64'].value = 'EF #6 VFD'
    sheet['A65'].value = 'Core Pump #1 VFD'
    sheet['A66'].value = 'Core Pump #2 VFD'
    sheet['A67'].value = 'HP LL- 3 Ok  (Fan is ok, If there\'s sweating of pipes check operation of HP)'
    sheet['A68'].value = 'Core Pump #2 (15 - 20 PSID)'
    sheet['A69'].value = 'Core Pump #1 (15 - 20 PSID)'
    sheet['A70'].value = 'Condenser Supply Temp.  West Side (68 – 85)'
    sheet['A71'].value = 'Chemical tanks level (above the order lines)'
    sheet['A72'].value = 'Nalco controller'
    sheet['A73'].value = 'Coupon Rack flow is between 4 – 6 GPM'
    sheet['A74'].value = 'Tower #4 VFD'
    sheet['A75'].value = 'Tower #3 VFD'
    sheet['A76'].value = 'Tower #2 VFD'
    sheet['A77'].value = 'Tower #1 VFD'
    sheet['A78'].value = 'Notes:' # StretchGoal: Increase row height, delete comment rows below
    print('04-05')
    wb.save('Plymouth_Daily_Rounds.xlsx')


def pg04_engineer_values():
    # Engineering Values
    # Local Variables
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='bottom')
    columnEven = [2, 4, 6, 8]
    columnOdd = [3, 5, 7, 9]

    # Yes or No values
    rows = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 22, 24]
    # cells = []
    for col in columnEven:
        for row in rows:
            sheet.cell(row=row, column=col).value = 'Yes  /  No'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size = 8, i=True, color='000000')
    # ✓ X values
    rowsCheck = [6, 7, 8, 9, 10, 15, 16, 17, 25, 26]
    for col in columnEven:
        for row in rowsCheck:
            # print(col, row)
            sheet.cell(row=row, column=col).value = '✓  or  X'
            sheet.cell(row=row, column=col).alignment = center
            sheet.cell(row=row, column=col).font = Font(size=9, color='DCDCDC')
    '''    # Hz
    rowsHZ = [18]
    for col in columnOdd:
        for row in rowsHZ:
            # print(col, row)
            sheet.cell(row=row, column=col).value = 'Hz'
            sheet.cell(row=row, column=col).alignment = right
            sheet.cell(row=row, column=col).font = Font(size=8, color='000000') '''
    print('04-06')
    wb.save('Plymouth_Daily_Rounds.xlsx')

def pg04_colored_cells():
    # Local Variables
    rowsColor = [1, 12, 24, 28]
    columnsColor = range(1, 10, 1)
    for col in columnsColor:
        for row in rowsColor:
            # print(col, row)
            sheet.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')
    print('04-07')
    wb.save('Plymouth_Daily_Rounds.xlsx')
