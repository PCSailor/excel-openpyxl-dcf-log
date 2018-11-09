#!/usr/bin/env python3
'''
* Looping practice
'''
print('\nStart next file, \'Loops\'')
# imports
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill
wb = load_workbook(filename = 'Plymouth_Daily_Rounds.xlsx')
sheet = wb["Loops"]
print('Active sheet is ', sheet)
wb.save('Plymouth_Daily_Rounds.xlsx')
print('Loops - test - 01')

values = ['Mechanical / Chill Water Units Room continued', 'DP 3 Breakers', 'DP3-B08 SPD All 3 Green lights (Protected)',  'DP3-B12 CHILLER 3 Breaker is',  'DP3-B13 ATS-HPC-C Breaker is', 'DP3-B14 ATS-HPC-H Breaker is', 'DP3-B15 MUA 4 Breaker is', 'DP3-B17 ATS-HPD-H Breaker is', 'Notes']

def loops_cell_values():
    values = ['Mechanical / Chill Water Units Room continued', 'DP 3 Breakers', 'DP3-B08 SPD All 3 Green lights (Protected)',  'DP3-B12 CHILLER 3 Breaker is',  'DP3-B13 ATS-HPC-C Breaker is', 'DP3-B14 ATS-HPC-H Breaker is', 'DP3-B15 MUA 4 Breaker is', 'DP3-B17 ATS-HPD-H Breaker is', 'Notes']
    print('Loops - test - 02')
    # Cell values
    for v in values:
        print(v)
    # Original Code from Python-forum.io:
    def update_vertically(sheet, col, row, values):
        row = int(row)
        for i, value in enumerate(values):
            sheet['{}{}'.format(col, row+i)].value = value
    update_vertically(sheet, 'A', 2, values)
    wb.save('Plymouth_Daily_Rounds.xlsx')


print('Loops - test - 03')

def pg09_cell_values():
    # Cell values
    values = ['STARTING HERE', 'Mechanical / Chill Water Units Room continued', 'DP 3 Breakers', 'DP3-B08 SPD All 3 Green lights (Protected)',  'DP3-B12 CHILLER 3 Breaker is',  'DP3-B13 ATS-HPC-C Breaker is', 'DP3-B14 ATS-HPC-H Breaker is', 'DP3-B15 MUA 4 Breaker is', 'DP3-B17 ATS-HPD-H Breaker is', 'Notes']
    print('Loops - test - 04')
    for v in values:
        print(v)
    def update_vertically(sheet, col, row, values):
        row = int(row)
        for i, value in enumerate(values):
            sheet['{}{}'.format(col, row+i)].value = value
    update_vertically(sheet, 'A', 2, values)
    wb.save('Plymouth_Daily_Rounds.xlsx')